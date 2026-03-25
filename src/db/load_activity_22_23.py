"""
Load activity survey data for 2022 and 2023 into ctis_stats.db.

Parses Excel WT/TAB tables for:
  a1: revenue, revenue_climate
  a2: revenue_no_reason
  a3: rnd_expense, rnd_expense_climate
  a4: employee_count
  a5: employee_count_climate
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
import re
import sqlite3
import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # src/
DB_PATH = os.path.join(BASE, 'db', 'ctis_stats.db')
DATA_DIR = os.path.join(BASE, 'data', '활동조사')

FILE_22 = os.path.join(DATA_DIR, '기후기술조사(WT테이블)0207 -22년도.xlsx')
FILE_23 = os.path.join(DATA_DIR, '2023년도 TAB_기후기술_V1_가중.xlsx')

# ── Category name mapping (Excel name → DB category_name) ─────────────
# DB T22 categories (taxonomy_id=3):
#   83: 태양광.열, 84: 풍력, 85: 해양 및 수력에너지, 86: 수열 및 지열,
#   87: 바이오, 88: 수소암모니아, 89: 비재생에너지, 90: 수소바이오매스,
#   91: 폐자원, 92: 발전효율, 93: 산업효율, 94: 수송효율, 95: 건물효율,
#   96: 포집/저장/흡수/대체, 97: 전력통합,
#   98: 기후변화 모니터링, 99: 기후영향평가 및 진단, 100: 건강, 101: 물,
#   102: 농축수산, 103: 기타 탄력성 강화, 104: 적응기반

def load_category_map(conn):
    """Return dict: normalized_name -> category_id for T22 taxonomy."""
    cur = conn.cursor()
    cur.execute("SELECT category_id, category_name FROM tech_category WHERE taxonomy_id=3")
    mapping = {}
    for cid, cname in cur.fetchall():
        mapping[normalize_name(cname)] = cid
    return mapping

def normalize_name(name):
    """Normalize category name for matching: strip prefixes, whitespace, punctuation quirks."""
    s = str(name).strip()
    # Remove (감축) / (적응) prefix
    s = re.sub(r'^\(감축\)', '', s)
    s = re.sub(r'^\(적응\)', '', s)
    # Remove leading number+dot like "01. "
    s = re.sub(r'^\d{1,2}\.\s*', '', s)
    # Replace middle dot · with .
    s = s.replace('·', '.')
    s = s.strip()
    return s


# ── Org-size mapping ──────────────────────────────────────────────────

ORG_SIZE_PATTERNS_22 = [
    (r'층1|100억 미만', '100억미만'),
    (r'층2|100억 이상 500억 미만', '100~500억'),
    (r'층3|500억 이상 1천억 미만', '500~1000억'),
    (r'층4|1천억 이상 2천억 미만', '1000~2000억'),
    (r'층5|2천억 이상', '2000억이상'),
    (r'매출액 정보 없음|공공기관 등', '공공기관등'),
]

ORG_SIZE_PATTERNS_23 = [
    (r'1\.\s*100억 미만', '100억미만'),
    (r'2\.\s*100억 이상', '100~500억'),
    (r'3\.\s*500억 이상', '500~1000억'),
    (r'4\.\s*1000억 이상', '1000~2000억'),
    (r'5\.\s*2000억 이상', '2000억이상'),
    (r'6\.\s*기타 공공부문|기타 공공부문 등', '공공기관등'),
]

def match_org_size(text, patterns):
    """Match org_size label to standardized name. Return None if no match."""
    if not text:
        return None
    s = str(text).strip()
    for pat, label in patterns:
        if re.search(pat, s):
            return label
    return None


# ── Excel table finder ─────────────────────────────────────────────────

def find_table_rows(ws, marker_text, marker_style='22'):
    """
    Find the start row of a table by its marker text.
    Returns the row number of the marker line, or None.
    marker_style: '22' for 【표 ...】 or '23' for ■ 표 ...
    """
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        val = str(cell.value) if cell.value else ''
        if marker_text in val:
            return cell.row
    return None


# ── Parse generic table (a1/a3/a4/a5 style) ───────────────────────────

def parse_table_22_simple(ws, marker_row, cat_map, value_col_index):
    """
    Parse a 22년도 table starting at marker_row.
    Structure:
      marker_row:   【표 ...】
      marker_row+1: Go
      marker_row+2: header row 1
      marker_row+3: header row 2
      marker_row+4: [전  체] → org_size='전체'
      marker_row+5: 적응/감축 감축 ... (subtotal, skip)
      marker_row+6: 적응 ... (subtotal, skip)
      marker_row+7..marker_row+7+21: category rows (22 categories)
      then org_size rows

    value_col_index: 1-based column index for the value to extract (합계).
    Returns list of (category_id_or_None, org_size, value).
    category_id=None means '전체' or org_size row.
    """
    results = []
    r = marker_row + 4  # [전  체] row

    # Read [전체] row
    total_val = ws.cell(r, value_col_index).value
    if total_val is not None and total_val != '.':
        results.append((None, '전체', _to_float(total_val)))

    r += 1  # Skip 감축 subtotal
    r += 1  # Skip 적응 subtotal

    # Skip additional subtotal row if present (적응/감축)
    # Actually row+5 is 적응/감축 | 감축, row+6 is None | 적응
    r += 1  # Now at first category row

    # Read category rows until we hit 기관규모
    while r <= marker_row + 40:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value

        # Detect end / org_size section
        if col_a and '기관규모' in str(col_a):
            break
        if col_a is None and col_b is None:
            break

        # Category name is in col_b (col_a is None for most, or '기후기술 분야' for first)
        cat_name_raw = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        if not cat_name_raw:
            r += 1
            continue

        norm = normalize_name(cat_name_raw)
        cat_id = cat_map.get(norm)
        if cat_id is None:
            print(f"  WARNING: No category match for '{cat_name_raw}' (normalized: '{norm}')")
            r += 1
            continue

        val = ws.cell(r, value_col_index).value
        if val is not None and val != '.':
            results.append((cat_id, '전체', _to_float(val)))
        r += 1

    # Now read org_size rows
    while r <= marker_row + 50:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value
        if col_a is None and col_b is None:
            break

        label = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        org = match_org_size(label, ORG_SIZE_PATTERNS_22)
        if org:
            val = ws.cell(r, value_col_index).value
            if val is not None and val != '.':
                results.append((None, org, _to_float(val)))
        r += 1

    return results


def parse_table_23_simple(ws, marker_row, cat_map, value_col_index):
    """
    Parse a 23년도 table starting at marker_row.
    Structure:
      marker_row:   ■ 표 ...
      marker_row+1: header row 1 (사례수, [합계], ...)
      marker_row+2: header row 2 (sub-headers)
      marker_row+3: [전체] → org_size='전체'
      marker_row+4: 적응/감축 | 감축 (subtotal)
      marker_row+5: None | 적응 (subtotal)
      marker_row+6...: 기후기술 분야(감축) | 01. 태양광.열 ...
      ...
      기후기술 분야(적응) | 16. 기후변화 모니터링 ...
      ...
      기관규모 | 1. 100억 미만 ...

    value_col_index: 1-based column index for the value (합계 = col 4).
    """
    results = []
    r = marker_row + 3  # [전체] row

    # Read [전체] row
    total_val = ws.cell(r, value_col_index).value
    if total_val is not None and total_val != '.':
        results.append((None, '전체', _to_float(total_val)))

    r += 1  # 감축 subtotal
    r += 1  # 적응 subtotal
    r += 1  # first category row

    # Read category rows
    while r <= marker_row + 40:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value

        if col_a and '기관규모' in str(col_a):
            break
        if col_a is None and col_b is None:
            break

        cat_name_raw = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        if not cat_name_raw:
            r += 1
            continue

        # Skip section headers like "기후기술 분야(감축)" or "기후기술 분야(적응)"
        if '기후기술 분야' in cat_name_raw and col_b is None:
            r += 1
            continue

        norm = normalize_name(cat_name_raw)
        cat_id = cat_map.get(norm)
        if cat_id is None:
            print(f"  WARNING: No category match for '{cat_name_raw}' (normalized: '{norm}')")
            r += 1
            continue

        val = ws.cell(r, value_col_index).value
        if val is not None and val != '.':
            results.append((cat_id, '전체', _to_float(val)))
        r += 1

    # Org-size rows
    while r <= marker_row + 50:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value
        if col_a is None and col_b is None:
            break

        label = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        org = match_org_size(label, ORG_SIZE_PATTERNS_23)
        if org:
            val = ws.cell(r, value_col_index).value
            if val is not None and val != '.':
                results.append((None, org, _to_float(val)))
        r += 1

    return results


# ── Parse A3-1 매출미발생 사유 table ──────────────────────────────────

def parse_a3_1_table_22(ws, marker_row, cat_map):
    """
    Parse 22년도 A3-1 매출미발생이유 table.
    Columns (1-based):
      1: group, 2: category, 3: 사례수,
      4-12: reason columns (사업화자금부족, 시장여건변화대응미흡, ...)
    Each row's values are percentages summing to ~100%.
    We store each reason as a separate metric_value entry with sub_metric in metric_type.

    Actually per spec: metric_type='revenue_no_reason', and we store each reason column.
    Let's store the full row of reasons as separate inserts.
    """
    reason_cols_22 = [
        (4, '사업화자금부족'),
        (5, '시장여건변화대응미흡'),
        (6, '제품서비스완성도미비'),
        (7, '판로개척실패'),
        (8, '추가기술개발실패'),
        (9, '사업화전문인력부족'),
        (10, '마케팅홍보역량부족'),
        (11, '법규제정보획득대응어려움'),
        (12, '기타'),
    ]
    return _parse_reason_table(ws, marker_row, cat_map, reason_cols_22,
                                ORG_SIZE_PATTERNS_22, skip_go_row=True)


def parse_a3_1_table_23(ws, marker_row, cat_map):
    """Parse 23년도 A3-1 매출미발생이유 table."""
    reason_cols_23 = [
        (4, '사업화자금부족'),
        (5, '시장여건변화대응미흡'),
        (6, '제품서비스완성도미비'),
        (7, '판로개척실패'),
        (8, '추가기술개발실패'),
        (9, '사업화전문인력부족'),
        (10, '마케팅홍보역량부족'),
        (11, '법규제정보획득대응어려움'),
        (12, '기타'),
    ]
    return _parse_reason_table(ws, marker_row, cat_map, reason_cols_23,
                                ORG_SIZE_PATTERNS_23, skip_go_row=False)


def _parse_reason_table(ws, marker_row, cat_map, reason_cols, org_patterns, skip_go_row):
    """
    Generic parser for 매출미발생이유 tables.
    Returns list of (category_id_or_None, org_size, reason_name, value).
    """
    results = []
    offset = 4 if skip_go_row else 3  # skip marker, Go/headers vs marker, headers

    r = marker_row + offset - 1  # [전체] row

    # Determine data start: find [전체] or [전  체]
    for scan_r in range(marker_row + 1, marker_row + 6):
        val = str(ws.cell(scan_r, 1).value or '')
        if '전' in val and '체' in val:
            r = scan_r
            break

    # Read [전체]
    for col_idx, reason_name in reason_cols:
        val = ws.cell(r, col_idx).value
        if val is not None and val != '.':
            results.append((None, '전체', reason_name, _to_float(val)))
    r += 1

    # Skip subtotals (감축, 적응)
    r += 1  # 감축
    r += 1  # 적응

    # Category rows
    while r <= marker_row + 45:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value

        if col_a and '기관규모' in str(col_a):
            break
        if col_a is None and col_b is None:
            break

        cat_name_raw = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        if not cat_name_raw:
            r += 1
            continue
        # Skip section headers
        if '기후기술 분야' in cat_name_raw and col_b is None:
            r += 1
            continue

        norm = normalize_name(cat_name_raw)
        cat_id = cat_map.get(norm)
        if cat_id is None:
            print(f"  WARNING (a2): No category match for '{cat_name_raw}' (normalized: '{norm}')")
            r += 1
            continue

        for col_idx, reason_name in reason_cols:
            val = ws.cell(r, col_idx).value
            if val is not None and val != '.':
                results.append((cat_id, '전체', reason_name, _to_float(val)))
        r += 1

    # Org-size rows
    while r <= marker_row + 55:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value
        if col_a is None and col_b is None:
            break

        label = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        org = match_org_size(label, org_patterns)
        if org:
            for col_idx, reason_name in reason_cols:
                val = ws.cell(r, col_idx).value
                if val is not None and val != '.':
                    results.append((None, org, reason_name, _to_float(val)))
        r += 1

    return results


def _to_float(val):
    """Safely convert value to float."""
    if val is None or val == '.' or val == '-':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ── Main ────────────────────────────────────────────────────────────────

def main():
    conn = sqlite3.connect(DB_PATH)
    cat_map = load_category_map(conn)
    print(f"Loaded {len(cat_map)} T22 categories from DB")
    print(f"Category map: { {k: v for k, v in sorted(cat_map.items(), key=lambda x: x[1])} }")

    # ── Step 1: Clear existing data ──
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM activity_survey")
    old_count = cur.fetchone()[0]
    cur.execute("DELETE FROM activity_survey")
    conn.commit()
    print(f"\nDeleted {old_count} existing rows from activity_survey")

    # ── Step 2: Load workbooks ──
    print(f"\nLoading 22년도: {os.path.basename(FILE_22)}")
    wb22 = openpyxl.load_workbook(FILE_22, data_only=True)
    ws22 = wb22['테이블']

    print(f"Loading 23년도: {os.path.basename(FILE_23)}")
    wb23 = openpyxl.load_workbook(FILE_23, data_only=True)
    ws23 = wb23['Table']

    all_inserts = []

    # ──────────────────────────────────────────────────────────────────
    # 2022 tables
    # ──────────────────────────────────────────────────────────────────
    print("\n=== Parsing 2022 data ===")

    # a1: revenue (A3_1_1 합계 col=5)
    row = find_table_rows(ws22, 'A3_1_1. 2022년 매출')
    if row:
        data = parse_table_22_simple(ws22, row, cat_map, value_col_index=5)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'revenue', val, '백만원', 1))
        print(f"  a1 revenue: {len(data)} rows from row {row}")

    # a1: revenue_climate (A3_1_4 합계 col=5)
    row = find_table_rows(ws22, 'A3_1_4. 2022년 기업 규모별 매출 현황_기후기술 분야 매출액')
    if row:
        data = parse_table_22_simple(ws22, row, cat_map, value_col_index=5)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'revenue_climate', val, '백만원', 1))
        print(f"  a1 revenue_climate: {len(data)} rows from row {row}")

    # a2: revenue_no_reason (A3-1)
    row = find_table_rows(ws22, 'A3-1. 매출 미발생')
    if row:
        data = parse_a3_1_table_22(ws22, row, cat_map)
        for cat_id, org, reason, val in data:
            mt = f'revenue_no_reason_{reason}'
            all_inserts.append((2022, cat_id, org, mt, val, '%', 1))
        print(f"  a2 revenue_no_reason: {len(data)} rows from row {row}")

    # a3: rnd_expense (B2_1_1 합계 col=5)
    row = find_table_rows(ws22, 'B2_1_1. 2022년 연구개발')
    if row:
        data = parse_table_22_simple(ws22, row, cat_map, value_col_index=5)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'rnd_expense', val, '백만원', 1))
        print(f"  a3 rnd_expense: {len(data)} rows from row {row}")

    # a3: rnd_expense_climate (B2_1_4 합계 col=5)
    row = find_table_rows(ws22, 'B2_1_4. 2022년 기업 규모별 연구개발 현황_기후기술 분야 연구개발비')
    if row:
        data = parse_table_22_simple(ws22, row, cat_map, value_col_index=5)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'rnd_expense_climate', val, '백만원', 1))
        print(f"  a3 rnd_expense_climate: {len(data)} rows from row {row}")

    # a4: employee_count (C1 전체 종사자수, 합계 col=5)
    row = find_table_rows(ws22, 'C1. 종업원 수_전체 종사자수')
    if row:
        data = parse_table_22_simple(ws22, row, cat_map, value_col_index=5)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'employee_count', val, '명', 1))
        print(f"  a4 employee_count: {len(data)} rows from row {row}")

    # a5: employee_count_climate (C1 기후기술 종사자, 합계 col=5)
    row = find_table_rows(ws22, 'C1. 종업원 수_기후기술 분야 종사자 수')
    if row:
        data = parse_table_22_simple(ws22, row, cat_map, value_col_index=5)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'employee_count_climate', val, '명', 1))
        print(f"  a5 employee_count_climate: {len(data)} rows from row {row}")

    # ──────────────────────────────────────────────────────────────────
    # 2023 tables
    # ──────────────────────────────────────────────────────────────────
    print("\n=== Parsing 2023 data ===")

    # a1: revenue (A3_1_1 합계 col=4)
    row = find_table_rows(ws23, 'A3_1_1. 전체 매출액(2023년')
    if row:
        data = parse_table_23_simple(ws23, row, cat_map, value_col_index=4)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'revenue', val, '백만원', 1))
        print(f"  a1 revenue: {len(data)} rows from row {row}")

    # a1: revenue_climate (A3_1_3 합계 col=4)
    row = find_table_rows(ws23, 'A3_1_3. 기후기술분야 매출액(2023년')
    if row:
        data = parse_table_23_simple(ws23, row, cat_map, value_col_index=4)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'revenue_climate', val, '백만원', 1))
        print(f"  a1 revenue_climate: {len(data)} rows from row {row}")

    # a2: revenue_no_reason (A3-1)
    row = find_table_rows(ws23, 'A3-1. 매출 미발생')
    if row:
        data = parse_a3_1_table_23(ws23, row, cat_map)
        for cat_id, org, reason, val in data:
            mt = f'revenue_no_reason_{reason}'
            all_inserts.append((2023, cat_id, org, mt, val, '%', 1))
        print(f"  a2 revenue_no_reason: {len(data)} rows from row {row}")

    # a3: rnd_expense (B2_1_1 합계 col=4)
    row = find_table_rows(ws23, 'B2_1_1. 전체 연구개발비 (2023년')
    if row:
        data = parse_table_23_simple(ws23, row, cat_map, value_col_index=4)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'rnd_expense', val, '백만원', 1))
        print(f"  a3 rnd_expense: {len(data)} rows from row {row}")

    # a3: rnd_expense_climate (B2_1_3 합계 col=4)
    row = find_table_rows(ws23, 'B2_1_3. 기후기술분야 연구개발비(2023년')
    if row:
        data = parse_table_23_simple(ws23, row, cat_map, value_col_index=4)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'rnd_expense_climate', val, '백만원', 1))
        print(f"  a3 rnd_expense_climate: {len(data)} rows from row {row}")

    # a4: employee_count (C1_1_1 합계 col=4)
    row = find_table_rows(ws23, 'C1_1_1. 전체 종사자 수(전체)')
    if row:
        data = parse_table_23_simple(ws23, row, cat_map, value_col_index=4)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'employee_count', val, '명', 1))
        print(f"  a4 employee_count: {len(data)} rows from row {row}")

    # a5: employee_count_climate (C1_2_1 합계 col=4)
    row = find_table_rows(ws23, 'C1_2_1. 기후기술 분야별 종업원 수(전체)')
    if row:
        data = parse_table_23_simple(ws23, row, cat_map, value_col_index=4)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'employee_count_climate', val, '명', 1))
        print(f"  a5 employee_count_climate: {len(data)} rows from row {row}")

    # ── Step 3: Insert all ──
    print(f"\n=== Inserting {len(all_inserts)} rows ===")
    cur.executemany(
        """INSERT INTO activity_survey
           (survey_year, category_id, org_size, metric_type, metric_value, metric_unit, is_weighted)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        all_inserts
    )
    conn.commit()

    # ── Step 4: Summary ──
    print("\n=== Summary ===")
    cur.execute("""
        SELECT survey_year, metric_type, COUNT(*), SUM(metric_value)
        FROM activity_survey
        GROUP BY survey_year, metric_type
        ORDER BY survey_year, metric_type
    """)
    for year, mt, cnt, total in cur.fetchall():
        print(f"  {year} | {mt:40s} | {cnt:5d} rows | total={total:,.2f}" if total else
              f"  {year} | {mt:40s} | {cnt:5d} rows | total=N/A")

    cur.execute("SELECT COUNT(*) FROM activity_survey")
    total = cur.fetchone()[0]
    print(f"\nTotal rows in activity_survey: {total}")

    conn.close()
    print("\nDone!")


if __name__ == '__main__':
    main()
