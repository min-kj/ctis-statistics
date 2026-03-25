#!/usr/bin/env python3
"""
CTis 기후기술통계 DB 초기화 스크립트
- 7개 테이블 생성 + 4개 뷰 생성
- Excel 데이터 적재
# 🐜 Scout: PLAN-020 DB 초기화 - 2026-03-20
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import re
import sqlite3
from pathlib import Path

import openpyxl

# ── 경로 설정 ──────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent.parent  # ctis-stats/
DATA_DIR = BASE_DIR / "src" / "data"
MAPPING_DIR = BASE_DIR / "src" / "mapping"
DB_PATH = BASE_DIR / "src" / "db" / "ctis_stats.db"

# Excel 파일 경로
EXCEL_2025_SUMMARY = DATA_DIR / "2차델파이_5개국_요약테이블.xlsx"
EXCEL_2020_DATA = DATA_DIR / "수준조사 데이터(2020)_250926.xlsx"
EXCEL_MAPPING = DATA_DIR / "기술연계표_맵핑표.xlsx"
EXCEL_RAWDATA_2025 = DATA_DIR / "2025 수준조사 Rawdata" / "2차_최종DATA_(3차 포함).xlsx"
EXCEL_ACTIVITY_2022 = DATA_DIR / "활동조사" / "기후기술조사(WT테이블)0207 -22년도.xlsx"


# ── DDL ────────────────────────────────────────────────
DDL_TABLES = """
CREATE TABLE IF NOT EXISTS taxonomy (
    taxonomy_id   INTEGER PRIMARY KEY,
    code          VARCHAR(20)  NOT NULL UNIQUE,
    name          VARCHAR(100) NOT NULL,
    category_count INTEGER,
    detail_count  INTEGER,
    effective_year INTEGER,
    source        VARCHAR(200),
    is_current    BOOLEAN DEFAULT 0
);

CREATE TABLE IF NOT EXISTS tech_category (
    category_id   INTEGER PRIMARY KEY,
    taxonomy_id   INTEGER NOT NULL REFERENCES taxonomy(taxonomy_id),
    category_no   VARCHAR(10),
    category_name VARCHAR(200) NOT NULL,
    type_code     VARCHAR(10),
    type_name     VARCHAR(20),
    definition    TEXT
);

CREATE TABLE IF NOT EXISTS tech_detail (
    detail_id     INTEGER PRIMARY KEY,
    category_id   INTEGER NOT NULL REFERENCES tech_category(category_id),
    detail_no     VARCHAR(10),
    detail_name   VARCHAR(300) NOT NULL,
    survey_year   INTEGER NOT NULL
);

CREATE TABLE IF NOT EXISTS taxonomy_mapping (
    mapping_id     INTEGER PRIMARY KEY,
    src_category_id INTEGER NOT NULL REFERENCES tech_category(category_id),
    tgt_category_id INTEGER NOT NULL REFERENCES tech_category(category_id),
    mapping_type   VARCHAR(20),
    mapping_status VARCHAR(20) DEFAULT 'confirmed',
    note           TEXT
);

CREATE TABLE IF NOT EXISTS survey_result (
    result_id     INTEGER PRIMARY KEY,
    survey_year   INTEGER NOT NULL,
    detail_id     INTEGER NOT NULL REFERENCES tech_detail(detail_id),
    country_code  VARCHAR(5) NOT NULL,
    tech_level    REAL,
    tech_gap      REAL,
    gap_unit      VARCHAR(5),
    tech_group    VARCHAR(10),
    basic_research    REAL,
    applied_research  REAL,
    rd_trend      VARCHAR(10),
    is_leading    BOOLEAN DEFAULT 0
);

CREATE UNIQUE INDEX IF NOT EXISTS idx_survey_result_unique
    ON survey_result(survey_year, detail_id, country_code);

CREATE TABLE IF NOT EXISTS survey_response_raw (
    response_id     INTEGER PRIMARY KEY,
    survey_year     INTEGER NOT NULL DEFAULT 2025,
    delphi_round    INTEGER,
    respondent_id   INTEGER,
    detail_id       INTEGER REFERENCES tech_detail(detail_id),
    country         VARCHAR(30),
    tech_group      INTEGER,
    tech_level_pct  REAL,
    tech_gap_month  REAL,
    basic_research  INTEGER,
    applied_research INTEGER,
    rd_trend        INTEGER
);

CREATE TABLE IF NOT EXISTS activity_survey (
    activity_id   INTEGER PRIMARY KEY,
    survey_year   INTEGER NOT NULL,
    category_id   INTEGER REFERENCES tech_category(category_id),
    org_size      VARCHAR(50),
    metric_type   VARCHAR(50),
    metric_value  REAL,
    metric_unit   VARCHAR(20),
    is_weighted   BOOLEAN DEFAULT 1
);
"""

DDL_VIEWS = """
CREATE VIEW IF NOT EXISTS v_category_summary AS
SELECT
    sr.survey_year,
    tc.category_name,
    tc.type_name,
    sr.country_code,
    AVG(sr.tech_level)       AS avg_level,
    AVG(sr.tech_gap)         AS avg_gap,
    COUNT(sr.detail_id)      AS detail_count,
    SUM(sr.is_leading)       AS leading_count
FROM survey_result sr
JOIN tech_detail td ON sr.detail_id = td.detail_id
JOIN tech_category tc ON td.category_id = tc.category_id
GROUP BY sr.survey_year, tc.category_name, tc.type_name, sr.country_code;

CREATE VIEW IF NOT EXISTS v_country_summary AS
SELECT
    survey_year,
    country_code,
    AVG(tech_level)  AS avg_level,
    AVG(tech_gap)    AS avg_gap,
    COUNT(*)         AS tech_count
FROM survey_result
GROUP BY survey_year, country_code;

CREATE VIEW IF NOT EXISTS v_type_summary AS
SELECT
    sr.survey_year,
    tc.type_name,
    sr.country_code,
    AVG(sr.tech_level)  AS avg_level,
    AVG(sr.tech_gap)    AS avg_gap,
    COUNT(*)            AS tech_count
FROM survey_result sr
JOIN tech_detail td ON sr.detail_id = td.detail_id
JOIN tech_category tc ON td.category_id = tc.category_id
GROUP BY sr.survey_year, tc.type_name, sr.country_code;

CREATE VIEW IF NOT EXISTS v_activity_summary AS
SELECT
    a.survey_year,
    tc.category_name,
    tc.type_name,
    a.metric_type,
    a.metric_unit,
    SUM(CASE WHEN a.org_size = '전체' THEN a.metric_value END) AS total_value,
    COUNT(DISTINCT a.org_size) AS size_count
FROM activity_survey a
JOIN tech_category tc ON a.category_id = tc.category_id
WHERE a.is_weighted = 1
GROUP BY a.survey_year, tc.category_name, tc.type_name, a.metric_type, a.metric_unit;
"""


def safe_float(val):
    """안전하게 float 변환"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ('', '-', '.', 'N/A', 'n/a', 'None'):
        return None
    try:
        return float(s.replace(',', ''))
    except (ValueError, TypeError):
        return None


def safe_int(val):
    """안전하게 int 변환"""
    f = safe_float(val)
    return int(f) if f is not None else None


def normalize_name(name):
    """이름 정규화: 번호 접두사 제거, 특수문자 통일, 공백 정규화"""
    if not name:
        return ''
    s = str(name).strip()
    # Remove leading number prefix like "01. ", "1.", "128.", etc.
    s = re.sub(r'^\d+[\.\s]+', '', s)
    # Normalize ALL middle-dot variants to a canonical form
    # U+00B7 (middle dot), U+0387 (greek ano teleia), U+30FB (katakana middle dot)
    # U+FF65 (halfwidth katakana middle dot), U+2027 (hyphenation point), ? (question mark)
    dot_chars = '\u00b7\u0387\u30fb\uff65\u2027'
    for c in dot_chars:
        s = s.replace(c, '·')
    s = s.replace('?', '·')
    # Normalize whitespace
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def normalize_for_match(name):
    """더 공격적인 정규화: 매칭용 (기술 접미사 제거 등)"""
    s = normalize_name(name)
    # Remove trailing '기술' suffix
    if s.endswith(' 기술'):
        s = s[:-3].strip()
    elif s.endswith('기술'):
        s = s[:-2].strip()
    return s


# ══════════════════════════════════════════════════════════
# Phase 1: 참조 테이블
# ══════════════════════════════════════════════════════════

def load_taxonomy(conn):
    """taxonomy 테이블 - 4개 분류체계"""
    rows = [
        ('T44', 'NIGT 44대 기후기술(2020)', 44, 185, 2020,
         'NIGT 기술수준조사 2020', 0),
        ('T38', '38대 기후기술(기후기술법)', 38, 157, 2025,
         '기후기술촉진법 시행규칙 고시 2022', 1),
        ('T22', '22대 승인통계 분류', 22, None, 2022,
         '국가승인통계 기술개발활동조사', 1),
        ('T100', '탄소중립 100대 핵심기술', 100, None, 2022,
         '탄소중립 100대 핵심기술(2024 개정)', 0),
    ]
    conn.executemany(
        "INSERT OR IGNORE INTO taxonomy "
        "(code, name, category_count, detail_count, effective_year, source, is_current) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        rows
    )
    print(f"  taxonomy: {len(rows)} rows inserted")


def get_taxonomy_id(conn, code):
    row = conn.execute(
        "SELECT taxonomy_id FROM taxonomy WHERE code = ?", (code,)
    ).fetchone()
    return row[0] if row else None


def load_tech_category_t38(conn):
    """T38 중분류 38개"""
    tid = get_taxonomy_id(conn, 'T38')
    wb = openpyxl.load_workbook(str(EXCEL_2025_SUMMARY), read_only=True, data_only=True)
    ws = wb['대분류별_통계']

    rows_inserted = 0
    for row in ws.iter_rows(min_row=2, max_row=39, values_only=True):
        cat_name_raw = row[0]
        if cat_name_raw is None:
            continue
        cat_name = str(cat_name_raw).strip()
        cat_no = cat_name.split('.')[0].strip() if '.' in cat_name else str(rows_inserted + 1).zfill(2)

        no_int = int(cat_no) if cat_no.isdigit() else 99
        if no_int <= 32:
            type_code, type_name = 'MIT', '감축'
        else:
            type_code, type_name = 'ADP', '적응'

        conn.execute(
            "INSERT INTO tech_category "
            "(taxonomy_id, category_no, category_name, type_code, type_name) "
            "VALUES (?, ?, ?, ?, ?)",
            (tid, cat_no, cat_name, type_code, type_name)
        )
        rows_inserted += 1

    wb.close()
    print(f"  tech_category (T38): {rows_inserted} rows inserted")


def load_tech_category_t44(conn):
    """T44 중분류 44개"""
    tid = get_taxonomy_id(conn, 'T44')
    wb = openpyxl.load_workbook(str(EXCEL_2020_DATA), read_only=True, data_only=True)
    ws = wb['소분류별_집계데이터']

    seen = {}
    cat_no_counter = 0
    for row in ws.iter_rows(min_row=2, max_row=45, values_only=True):
        type_name = str(row[0]).strip() if row[0] else ''
        cat_name = str(row[1]).strip() if row[1] else ''
        if not cat_name or cat_name in seen:
            continue

        cat_no_counter += 1
        cat_no = str(cat_no_counter).zfill(2)
        type_code = 'MIT' if type_name == '감축' else 'ADP'
        seen[cat_name] = cat_no_counter

        conn.execute(
            "INSERT INTO tech_category "
            "(taxonomy_id, category_no, category_name, type_code, type_name) "
            "VALUES (?, ?, ?, ?, ?)",
            (tid, cat_no, cat_name, type_code, type_name)
        )

    wb.close()
    print(f"  tech_category (T44): {len(seen)} rows inserted")


def load_tech_category_t22(conn):
    """T22 중분류 22개"""
    tid = get_taxonomy_id(conn, 'T22')

    categories = [
        ('01', '태양광.열', 'MIT', '감축'),
        ('02', '풍력', 'MIT', '감축'),
        ('03', '해양 및 수력에너지', 'MIT', '감축'),
        ('04', '수열 및 지열', 'MIT', '감축'),
        ('05', '바이오', 'MIT', '감축'),
        ('06', '수소암모니아', 'MIT', '감축'),
        ('07', '비재생에너지', 'MIT', '감축'),
        ('08', '수소바이오매스', 'MIT', '감축'),
        ('09', '폐자원', 'MIT', '감축'),
        ('10', '발전효율', 'MIT', '감축'),
        ('11', '산업효율', 'MIT', '감축'),
        ('12', '수송효율', 'MIT', '감축'),
        ('13', '건물효율', 'MIT', '감축'),
        ('14', '포집/저장/흡수/대체', 'MIT', '감축'),
        ('15', '전력통합', 'MIT', '감축'),
        ('16', '기후변화 모니터링', 'ADP', '적응'),
        ('17', '기후영향평가 및 진단', 'ADP', '적응'),
        ('18', '건강', 'ADP', '적응'),
        ('19', '물', 'ADP', '적응'),
        ('20', '농축수산', 'ADP', '적응'),
        ('21', '기타 탄력성 강화', 'ADP', '적응'),
        ('22', '적응기반', 'ADP', '적응'),
    ]

    for cat_no, cat_name, type_code, type_name in categories:
        conn.execute(
            "INSERT INTO tech_category "
            "(taxonomy_id, category_no, category_name, type_code, type_name) "
            "VALUES (?, ?, ?, ?, ?)",
            (tid, cat_no, cat_name, type_code, type_name)
        )
    print(f"  tech_category (T22): {len(categories)} rows inserted")


def _build_category_lookup(conn, taxonomy_code):
    """category_name -> category_id lookup dict (with normalized key variant)"""
    tid = get_taxonomy_id(conn, taxonomy_code)
    rows = conn.execute(
        "SELECT category_id, category_name FROM tech_category WHERE taxonomy_id = ?",
        (tid,)
    ).fetchall()
    lookup = {}
    for cid, name in rows:
        lookup[name] = cid
        # Also store normalized version
        nname = normalize_name(name)
        if nname not in lookup:
            lookup[nname] = cid
    return lookup


def _find_category_id(cat_lookup, name):
    """Try exact, then normalized matching"""
    if name in cat_lookup:
        return cat_lookup[name]
    nname = normalize_name(name)
    if nname in cat_lookup:
        return cat_lookup[nname]
    # Partial match
    for cn, cid in cat_lookup.items():
        ncn = normalize_name(cn)
        if ncn and nname and (ncn in nname or nname in ncn):
            return cid
    return None


def load_tech_detail_2025(conn):
    """2025 세부기술 157개"""
    cat_lookup = _build_category_lookup(conn, 'T38')
    wb = openpyxl.load_workbook(str(EXCEL_2025_SUMMARY), read_only=True, data_only=True)
    ws = wb['5개국_요약']

    rows_inserted = 0
    detail_no_counter = 0
    for row in ws.iter_rows(min_row=2, max_row=158, values_only=True):
        cat_name_raw = str(row[0]).strip() if row[0] else ''
        detail_name = str(row[1]).strip() if row[1] else ''
        if not detail_name:
            continue

        category_id = _find_category_id(cat_lookup, cat_name_raw)
        if category_id is None:
            print(f"    WARNING: No T38 category for '{cat_name_raw}', skipping '{detail_name}'")
            continue

        detail_no_counter += 1
        detail_no_str = detail_name.split('.')[0] if '.' in detail_name and detail_name.split('.')[0].strip().isdigit() else str(detail_no_counter)

        conn.execute(
            "INSERT INTO tech_detail "
            "(category_id, detail_no, detail_name, survey_year) "
            "VALUES (?, ?, ?, 2025)",
            (category_id, detail_no_str, detail_name)
        )
        rows_inserted += 1

    wb.close()
    print(f"  tech_detail (2025): {rows_inserted} rows inserted")


def load_tech_detail_2020(conn):
    """2020 세부기술 185개"""
    cat_lookup = _build_category_lookup(conn, 'T44')
    wb = openpyxl.load_workbook(str(EXCEL_2020_DATA), read_only=True, data_only=True)
    ws = wb['세부기술별_데이터']

    rows_inserted = 0
    for row in ws.iter_rows(min_row=2, max_row=186, values_only=True):
        subcategory = str(row[3]).strip() if row[3] else ''  # 소분류 = T44
        detail_no = str(row[4]).strip() if row[4] else ''
        detail_name = str(row[5]).strip() if row[5] else ''
        if not detail_name:
            continue

        category_id = _find_category_id(cat_lookup, subcategory)
        if category_id is None:
            print(f"    WARNING: No T44 category for '{subcategory}', skipping '{detail_name}'")
            continue

        conn.execute(
            "INSERT INTO tech_detail "
            "(category_id, detail_no, detail_name, survey_year) "
            "VALUES (?, ?, ?, 2020)",
            (category_id, detail_no, detail_name)
        )
        rows_inserted += 1

    wb.close()
    print(f"  tech_detail (2020): {rows_inserted} rows inserted")


def load_taxonomy_mapping(conn):
    """44대↔38대 매핑"""
    cat_lookup_t44 = _build_category_lookup(conn, 'T44')
    cat_lookup_t38 = _build_category_lookup(conn, 'T38')

    wb = openpyxl.load_workbook(str(EXCEL_MAPPING), read_only=True, data_only=True)
    ws = wb['2020 세부185-44']

    pairs = {}
    for row in ws.iter_rows(min_row=2, max_row=186, values_only=True):
        cat44_name = str(row[1]).strip() if row[1] else ''
        cat38_name = str(row[2]).strip() if row[2] else ''
        if not cat44_name or not cat38_name or cat38_name == 'N/A':
            continue
        key = (cat44_name, cat38_name)
        pairs[key] = pairs.get(key, 0) + 1

    rows_inserted = 0
    for (cat44_name, cat38_name), detail_count in pairs.items():
        src_id = _find_category_id(cat_lookup_t44, cat44_name)
        tgt_id = _find_category_id(cat_lookup_t38, cat38_name)

        if src_id is None:
            print(f"    WARNING mapping: T44 '{cat44_name}' not found")
            continue
        if tgt_id is None:
            print(f"    WARNING mapping: T38 '{cat38_name}' not found")
            continue

        t44_targets = sum(1 for (s, _) in pairs if s == cat44_name)
        t38_sources = sum(1 for (_, t) in pairs if t == cat38_name)

        if t44_targets == 1 and t38_sources == 1:
            mapping_type = '1:1'
        elif t44_targets > 1:
            mapping_type = '1:N'
        elif t38_sources > 1:
            mapping_type = 'N:1'
        else:
            mapping_type = '1:1'

        conn.execute(
            "INSERT INTO taxonomy_mapping "
            "(src_category_id, tgt_category_id, mapping_type, mapping_status, note) "
            "VALUES (?, ?, ?, 'confirmed', ?)",
            (src_id, tgt_id, mapping_type,
             f"{cat44_name} -> {cat38_name} ({detail_count} details)")
        )
        rows_inserted += 1

    wb.close()
    print(f"  taxonomy_mapping: {rows_inserted} rows inserted")


# ══════════════════════════════════════════════════════════
# Phase 2: 수준조사 데이터
# ══════════════════════════════════════════════════════════

def _build_detail_lookup(conn, survey_year):
    """detail_name -> detail_id, with multiple normalized variants for matching"""
    rows = conn.execute(
        "SELECT detail_id, detail_name FROM tech_detail WHERE survey_year = ?",
        (survey_year,)
    ).fetchall()
    lookup = {}
    for did, name in rows:
        lookup[name] = did
        # Normalized (no number prefix, unified dots)
        nname = normalize_name(name)
        if nname not in lookup:
            lookup[nname] = did
        # Aggressive (also strip 기술 suffix)
        mname = normalize_for_match(name)
        if mname not in lookup:
            lookup[mname] = did
    return lookup


def _find_detail_id(detail_lookup, raw_name):
    """Try multiple strategies to match a detail name"""
    # 1. Exact
    if raw_name in detail_lookup:
        return detail_lookup[raw_name]
    # 2. Normalized
    nname = normalize_name(raw_name)
    if nname in detail_lookup:
        return detail_lookup[nname]
    # 3. Aggressive (strip 기술)
    mname = normalize_for_match(raw_name)
    if mname in detail_lookup:
        return detail_lookup[mname]
    # 4. Try adding/removing parenthetical suffixes
    # e.g., "생물학적 전환 기술" -> stored as "생물학적 전환(CO2)"
    # or "생물학적 전환 기술2" -> "생물학적 전환(CH4)"
    base = mname.rstrip('0123456789').strip()
    if base != mname and base in detail_lookup:
        return detail_lookup[base]
    # 5. Partial match on base
    for key, did in detail_lookup.items():
        nkey = normalize_for_match(key)
        if nkey and base and len(base) > 4:
            if nkey.startswith(base) or base.startswith(nkey):
                return did
    return None


COUNTRY_MAP_KR = {
    '한국': 'KR', '중국': 'CN', '일본': 'JP', '미국': 'US', 'EU': 'EU',
    '미국•EU': 'US',
}

COUNTRIES_2025 = ['KR', 'CN', 'JP', 'US', 'EU']


def load_survey_result_2025(conn):
    """2025 수준조사 결과 - 157 x 5 = 785건"""
    detail_lookup = _build_detail_lookup(conn, 2025)
    wb = openpyxl.load_workbook(str(EXCEL_2025_SUMMARY), read_only=True, data_only=True)
    ws = wb['5개국_요약']

    level_cols = {'KR': 3, 'CN': 4, 'JP': 5, 'US': 6, 'EU': 7}
    gap_cols = {'KR': 8, 'CN': 9, 'JP': 10, 'US': 11, 'EU': 12}

    rows_inserted = 0
    for row in ws.iter_rows(min_row=2, max_row=158, values_only=True):
        detail_name = str(row[1]).strip() if row[1] else ''
        if not detail_name:
            continue

        detail_id = detail_lookup.get(detail_name)
        if detail_id is None:
            detail_id = detail_lookup.get(normalize_name(detail_name))
        if detail_id is None:
            print(f"    WARNING 2025 survey: detail '{detail_name}' not found")
            continue

        leading_raw = str(row[2]).strip() if row[2] else ''
        leading_code = COUNTRY_MAP_KR.get(leading_raw, '')

        for cc in COUNTRIES_2025:
            tech_level = safe_float(row[level_cols[cc]])
            tech_gap = safe_float(row[gap_cols[cc]])
            is_leading = 1 if cc == leading_code else 0

            conn.execute(
                "INSERT OR IGNORE INTO survey_result "
                "(survey_year, detail_id, country_code, tech_level, tech_gap, "
                " gap_unit, is_leading) "
                "VALUES (2025, ?, ?, ?, ?, 'month', ?)",
                (detail_id, cc, tech_level, tech_gap, is_leading)
            )
            rows_inserted += 1

    wb.close()
    print(f"  survey_result (2025): {rows_inserted} rows inserted")


def load_survey_result_2020(conn):
    """2020 수준조사 결과 - 185 x 5 = 925건"""
    detail_lookup = _build_detail_lookup(conn, 2020)
    wb = openpyxl.load_workbook(str(EXCEL_2020_DATA), read_only=True, data_only=True)
    ws = wb['세부기술별_데이터']

    country_cols = {
        'KR': {'level': 7,  'gap': 8,  'group': 9,  'rd_trend': 22, 'basic': 23, 'applied': 24},
        'CN': {'level': 10, 'gap': 11, 'group': 12, 'rd_trend': 25, 'basic': 26, 'applied': 27},
        'JP': {'level': 13, 'gap': 14, 'group': 15, 'rd_trend': 28, 'basic': 29, 'applied': 30},
        'US': {'level': 16, 'gap': 17, 'group': 18, 'rd_trend': 31, 'basic': 32, 'applied': 33},
        'EU': {'level': 19, 'gap': 20, 'group': 21, 'rd_trend': 34, 'basic': 35, 'applied': 36},
    }

    rows_inserted = 0
    for row in ws.iter_rows(min_row=2, max_row=186, values_only=True):
        detail_name = str(row[5]).strip() if row[5] else ''
        if not detail_name:
            continue

        detail_id = detail_lookup.get(detail_name)
        if detail_id is None:
            detail_id = detail_lookup.get(normalize_name(detail_name))
        if detail_id is None:
            print(f"    WARNING 2020 survey: detail '{detail_name}' not found")
            continue

        leading_raw = str(row[6]).strip() if row[6] else ''
        leading_code = COUNTRY_MAP_KR.get(leading_raw, '')

        for cc, cols in country_cols.items():
            tech_level = safe_float(row[cols['level']])
            tech_gap = safe_float(row[cols['gap']])
            tech_group = str(row[cols['group']]).strip() if row[cols['group']] else None
            rd_trend = str(row[cols['rd_trend']]).strip() if row[cols['rd_trend']] else None
            basic_research = safe_float(row[cols['basic']])
            applied_research = safe_float(row[cols['applied']])
            is_leading = 1 if cc == leading_code else 0

            conn.execute(
                "INSERT OR IGNORE INTO survey_result "
                "(survey_year, detail_id, country_code, tech_level, tech_gap, "
                " gap_unit, tech_group, basic_research, applied_research, "
                " rd_trend, is_leading) "
                "VALUES (2020, ?, ?, ?, ?, 'year', ?, ?, ?, ?, ?)",
                (detail_id, cc, tech_level, tech_gap,
                 tech_group, basic_research, applied_research, rd_trend, is_leading)
            )
            rows_inserted += 1

    wb.close()
    print(f"  survey_result (2020): {rows_inserted} rows inserted")


def load_survey_response_raw_2025(conn):
    """2025 로데이터 - 14,856건
    Rawdata has plain detail names (no number prefix), so we use normalized lookup.
    """
    detail_lookup = _build_detail_lookup(conn, 2025)
    wb = openpyxl.load_workbook(str(EXCEL_RAWDATA_2025), read_only=True, data_only=True)
    ws = wb['2차 데이터']

    # Columns: 0=사용자ID, 1=사용자명, 2=대분류기술분야, 3=세부기술분야,
    #          4=국가, 5=2차그룹, 6=2차비율, 7=2차격차(개월),
    #          8=기초연구, 9=응용개발연구, 10=연구개발동향

    rows_inserted = 0
    skipped = 0
    unmatched_names = set()

    for row in ws.iter_rows(min_row=2, max_row=14857, values_only=True):
        respondent_id = safe_int(row[0])
        detail_name_raw = str(row[3]).strip() if row[3] else ''
        country = str(row[4]).strip() if row[4] else ''

        detail_id = _find_detail_id(detail_lookup, detail_name_raw)

        if detail_id is None:
            skipped += 1
            if detail_name_raw not in unmatched_names:
                unmatched_names.add(detail_name_raw)
            continue

        tech_group = safe_int(row[5])
        tech_level_pct = safe_float(row[6])
        tech_gap_month = safe_float(row[7])
        basic_research = safe_int(row[8])
        applied_research = safe_int(row[9])
        rd_trend = safe_int(row[10])

        conn.execute(
            "INSERT INTO survey_response_raw "
            "(survey_year, delphi_round, respondent_id, detail_id, country, "
            " tech_group, tech_level_pct, tech_gap_month, "
            " basic_research, applied_research, rd_trend) "
            "VALUES (2025, 2, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (respondent_id, detail_id, country,
             tech_group, tech_level_pct, tech_gap_month,
             basic_research, applied_research, rd_trend)
        )
        rows_inserted += 1

    wb.close()
    if skipped:
        print(f"    (skipped {skipped} rows, {len(unmatched_names)} unique unmatched names)")
        if unmatched_names:
            for n in sorted(unmatched_names)[:10]:
                print(f"      unmatched: [{n}]")
    print(f"  survey_response_raw (2025): {rows_inserted} rows inserted")


# ══════════════════════════════════════════════════════════
# Phase 3: 활동조사
# ══════════════════════════════════════════════════════════

ACTIVITY_CATEGORY_PATTERNS = {
    '태양광.열': '(감축)태양광.열',
    '풍력': '(감축)풍력',
    '해양 및 수력에너지': '(감축)해양 및 수력에너지',
    '수열 및 지열': '(감축)수열 및 지열',
    '바이오': '(감축)바이오',
    '수소암모니아': '(감축)수소암모니아',
    '비재생에너지': '(감축)비재생에너지',
    '수소바이오매스': '(감축)수소바이오매스',
    '폐자원': '(감축)폐자원',
    '발전효율': '(감축)발전효율',
    '산업효율': '(감축)산업효율',
    '수송효율': '(감축)수송효율',
    '건물효율': '(감축)건물효율',
    '포집/저장/흡수/대체': '(감축)포집/저장/흡수/대체',
    '전력통합': '(감축)전력통합',
    '기후변화 모니터링': '(적응)기후변화 모니터링',
    '기후영향평가 및 진단': '(적응)기후영향평가 및 진단',
    '건강': '(적응)건강',
    '물': '(적응)물',
    '농축수산': '(적응)농축수산',
    '기타 탄력성 강화': '(적응)기타 탄력성 강화',
    '적응기반': '(적응)적응기반',
}

# Reverse: WT label -> T22 category name
WT_LABEL_TO_T22 = {v: k for k, v in ACTIVITY_CATEGORY_PATTERNS.items()}


def _parse_activity_table_v2(ws, start_row, end_row, metric_type, metric_unit,
                              survey_year, cat_lookup, conn):
    """
    Parse one WT table. Structure (columns 0-based):
      Row+0: 【표 ...】
      Row+1: 'Go'
      Row+2: header [c2=[사례수], c3=전체label, c5=층1label, c7=층2, c9=층3, c11=층4, c13=층5]
      Row+3: subheader [c3=평균, c4=합계, c5=평균, c6=합계, ...]
      Row+4+: data

    Category rows: c0 or c1 has label like '(감축)태양광.열', c2 has (count)
    Data: c3=전체평균, c4=전체합계, c5=층1평균, c6=층1합계, ...
    """
    # org_size -> 합계 column index
    ORG_SUM_COLS = {
        '전체': 4,
        '100억미만': 6,
        '100~500억': 8,
        '500~1000억': 10,
        '1000~2000억': 12,
        '2000억이상': 14,
    }

    rows_inserted = 0
    all_rows = list(ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True))

    for row_data in all_rows[4:]:  # skip title, 'Go', header, subheader
        if row_data is None:
            continue

        c0 = str(row_data[0]).strip() if row_data[0] else ''
        c1 = str(row_data[1]).strip() if row_data[1] else ''

        # Determine category label
        label = None
        if c1.startswith('(감축)') or c1.startswith('(적응)'):
            label = c1
        elif c0.startswith('(감축)') or c0.startswith('(적응)'):
            label = c0
        else:
            continue

        # Skip section headers / subtotals / org_size rows
        if label.startswith('기후기술') or label.startswith('기관규모'):
            continue

        # Match to T22 category
        t22_name = WT_LABEL_TO_T22.get(label)
        if t22_name is None:
            continue

        category_id = cat_lookup.get(t22_name)
        if category_id is None:
            continue

        # Extract values for each org_size
        for org_size, col_idx in ORG_SUM_COLS.items():
            try:
                val = safe_float(row_data[col_idx])
            except (IndexError, TypeError):
                val = None
            if val is None:
                continue

            conn.execute(
                "INSERT INTO activity_survey "
                "(survey_year, category_id, org_size, metric_type, "
                " metric_value, metric_unit, is_weighted) "
                "VALUES (?, ?, ?, ?, ?, ?, 1)",
                (survey_year, category_id, org_size, metric_type, val, metric_unit)
            )
            rows_inserted += 1

    return rows_inserted


def load_activity_survey_2022(conn):
    """활동조사 2022 - WT테이블"""
    cat_lookup = _build_category_lookup(conn, 'T22')
    wb = openpyxl.load_workbook(str(EXCEL_ACTIVITY_2022), read_only=True, data_only=True)
    ws = wb['테이블']
    max_row = ws.max_row

    # Find table markers
    table_starts = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
        first = str(row[0]).strip() if row[0] else ''
        if first.startswith('【표'):
            table_starts.append((i, first))

    # Build ranges
    tables = []
    for idx, (start, title) in enumerate(table_starts):
        end = table_starts[idx + 1][0] - 1 if idx + 1 < len(table_starts) else max_row
        tables.append((start, end, title))

    target_tables = {
        'A3_1_3': ('revenue', '백만원'),
        'A3_1_4': ('revenue_climate', '백만원'),
        'B2_1_3': ('rnd_expense', '백만원'),
        'B2_1_4': ('rnd_expense_climate', '백만원'),
        'C1. 종업원 수_전체': ('employee_count', '명'),
        'C1. 종업원 수_기후기술': ('employee_count_climate', '명'),
    }

    total_inserted = 0
    for start, end, title in tables:
        for key, (metric_type, metric_unit) in target_tables.items():
            if key in title:
                n = _parse_activity_table_v2(
                    ws, start, end, metric_type, metric_unit, 2022,
                    cat_lookup, conn
                )
                total_inserted += n
                print(f"    {title[:60]}: {n} rows")
                break

    wb.close()
    print(f"  activity_survey (2022): {total_inserted} rows inserted")


# ══════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════

def print_summary(conn):
    print("\n" + "=" * 60)
    print("  DB Summary")
    print("=" * 60)

    tables = ['taxonomy', 'tech_category', 'tech_detail', 'taxonomy_mapping',
              'survey_result', 'survey_response_raw', 'activity_survey']
    for tbl in tables:
        cnt = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        print(f"  {tbl:25s}: {cnt:>8,} rows")

    print("\n  --- Views ---")

    print("\n  v_country_summary:")
    for row in conn.execute(
        "SELECT survey_year, country_code, "
        "ROUND(avg_level, 2), ROUND(avg_gap, 2), tech_count "
        "FROM v_country_summary ORDER BY survey_year, country_code"
    ).fetchall():
        print(f"    {row}")

    print("\n  v_type_summary (2025):")
    for row in conn.execute(
        "SELECT survey_year, type_name, country_code, "
        "ROUND(avg_level, 2), ROUND(avg_gap, 2), tech_count "
        "FROM v_type_summary WHERE survey_year = 2025 "
        "ORDER BY type_name, country_code"
    ).fetchall():
        print(f"    {row}")

    print("\n  v_activity_summary:")
    for row in conn.execute(
        "SELECT survey_year, category_name, metric_type, "
        "ROUND(total_value, 0), metric_unit "
        "FROM v_activity_summary "
        "WHERE total_value IS NOT NULL "
        "ORDER BY metric_type, category_name "
        "LIMIT 15"
    ).fetchall():
        print(f"    {row}")


def main():
    print(f"CTis DB Init")
    print(f"  DB: {DB_PATH}")
    print(f"  Data: {DATA_DIR}")
    print()

    if DB_PATH.exists():
        DB_PATH.unlink()
        print("  Removed existing DB")

    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    try:
        print("Creating tables...")
        conn.executescript(DDL_TABLES)

        # Phase 1
        print("\nPhase 1: Reference tables")
        conn.execute("BEGIN")
        load_taxonomy(conn)
        load_tech_category_t38(conn)
        load_tech_category_t44(conn)
        load_tech_category_t22(conn)
        load_tech_detail_2025(conn)
        load_tech_detail_2020(conn)
        load_taxonomy_mapping(conn)
        conn.commit()

        # Phase 2
        print("\nPhase 2: Survey data")
        conn.execute("BEGIN")
        load_survey_result_2025(conn)
        load_survey_result_2020(conn)
        load_survey_response_raw_2025(conn)
        conn.commit()

        # Phase 3
        print("\nPhase 3: Activity survey")
        conn.execute("BEGIN")
        load_activity_survey_2022(conn)
        conn.commit()

        # Views
        print("\nCreating views...")
        conn.executescript(DDL_VIEWS)

        print_summary(conn)

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        conn.close()

    print(f"\nDone. DB: {DB_PATH}")
    print(f"  Size: {DB_PATH.stat().st_size / 1024:.1f} KB")


if __name__ == '__main__':
    main()
