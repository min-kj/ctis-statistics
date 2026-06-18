"""
2025 기후기술수준조사 세부기술별 데이터 생성 스크립트

소스:
  - ★보고서버전_최종.xlsx (L3) → 분류체계, 기술수준(100환산), 격차(0년환산)
  - 2차_최종DATA_(3차 포함).xlsx (L0) → 연구역량, R&D활동
    * 한국/중국/일본/미국: 해당국 개인응답 평균
    * EU: EU 회원국 중 국가별 평균의 최고국 값

출력: 기후기술수준조사_세부기술별 데이터(2025).xlsx
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
from collections import defaultdict, Counter
import openpyxl
from copy import copy

DATA = Path("src/data")

def _find(pattern):
    results = list(DATA.glob(pattern))
    if not results:
        raise FileNotFoundError(f"'{pattern}' not found in {DATA}")
    return results[0]

REPORT_XLSX = _find("*보고서버전_최종*")
L0_RAW = _find("2025 수준조사 Rawdata/*2차_최종DATA*")
TEMPLATE = _find("*세부기술별*2020*")
OUT = DATA / TEMPLATE.name.replace("2020", "2025")

COUNTRIES = ['한국', '중국', '일본', '미국', 'EU']
EU_MEMBER_COUNTRIES = {
    '독일', '프랑스', '영국', '이탈리아', '스페인', '네덜란드', '벨기에',
    '오스트리아', '스웨덴', '덴마크', '핀란드', '아일랜드', '노르웨이', '스위스',
    '룩셈부르크', '폴란드', '체코', '포르투갈', '에스토니아',
}
RAW_SCORE_MAP = {1: 20, 2: 40, 3: 60, 4: 80, 5: 100}
RAW_TREND_MAP = {1: '급하강', 2: '하강', 3: '유지', 4: '상승', 5: '급상승'}
TREND_ORDER = {'급상승': 5, '상승': 4, '유지': 3, '하강': 2, '급하강': 1}

# 38대 표준 대분류명 (플랫폼탑재용 기준)
STD_38_NAMES = {
    '01': '태양광 기술',       '02': '태양열 기술',       '03': '풍력 기술',
    '04': '해양에너지 기술',    '05': '수력 기술',        '06': '수열 기술',
    '07': '지열 기술',         '08': '바이오에너지 기술',   '09': '수소·암모니아 기술',
    '10': '석탄액화·가스화 기술', '11': '원자력 기술',       '12': '핵융합에너지 기술',
    '13': '수소 기술',         '14': '바이오매스 기술',    '15': '폐자원 기술',
    '16': '발전효율 기술',      '17': '산업효율 기술',      '18': '수송효율 기술',
    '19': '건물효율 기술',
    '20': '이산화탄소(CO2)포집, 저장, 활용 기술',
    '21': '메탄(CH)처리 기술',
    '22': '기타 온실가스 처리 및 대체 기술',
    '23': '탄소흡수원 기술',    '24': '전력 통합 기술',     '25': '열 통합 기술',
    '26': '전력-비전력 부문간 결합 기술',
    '27': '기후변화 감시 및 진단 기술',  '28': '기후변화 예측 기술',
    '29': '기후변화 영향 평가 기술',
    '30': '기후변화 취약성 및 위험성 평가 기술',
    '31': '건강부문 기술',      '32': '물 부문 기술',       '33': '국토연안 부문 기술',
    '34': '농축수산 부문 기술',   '35': '산림·생태계 부문 기술', '36': '산업·에너지 부문 기술',
    '37': '적응조치의 효과평가 기술', '38': '기후변화 적응기반 기술',
}

UPPER_TO_CLASS = {
    'I. 에너지 생산': '감축', 'II. 연원료 대체': '감축',
    'III. 에너지 효율': '감축', 'IV. 온실가스 처리': '감축',
    'V. 에너지 융복합': '감축',
    'Ⅵ. 기후변화 모니터링': '적응', 'Ⅶ. 기후영향 평가 및 진단': '적응',
    'Ⅷ. 피해관리 및 탄력성 제고': '적응', 'Ⅸ. 정책기술 분석 및 평가': '적응',
}


def _normalize_name(name):
    if not name:
        return name
    for ch in ['?', '\uff65', '\u30fb']:
        name = name.replace(ch, '\u00b7')
    return name.strip()


def _calc_research(records):
    """개인응답 리스트 → (기초점수, 응용점수, R&D경향)"""
    basics = [RAW_SCORE_MAP[d['basic']] for d in records if d.get('basic') in RAW_SCORE_MAP]
    apps = [RAW_SCORE_MAP[d['app']] for d in records if d.get('app') in RAW_SCORE_MAP]
    trends = [d['trend'] for d in records if d.get('trend') in RAW_TREND_MAP]
    avg_b = round(sum(basics) / len(basics), 1) if basics else None
    avg_a = round(sum(apps) / len(apps), 1) if apps else None
    rd = RAW_TREND_MAP[Counter(trends).most_common(1)[0][0]] if trends else None
    return avg_b, avg_a, rd


def _calc_eu(country_data):
    """EU 회원국별 평균 → 각 지표별 최고국 값"""
    eu_avgs = {}
    for cn in EU_MEMBER_COUNTRIES:
        recs = country_data.get(cn, [])
        if not recs:
            continue
        b, a, t = _calc_research(recs)
        eu_avgs[cn] = {'basic': b, 'app': a, 'trend': t}
    if not eu_avgs:
        return None, None, None
    best_b = max((v['basic'] for v in eu_avgs.values() if v['basic']), default=None)
    best_a = max((v['app'] for v in eu_avgs.values() if v['app']), default=None)
    best_t = max((v['trend'] for v in eu_avgs.values() if v['trend']),
                 key=lambda x: TREND_ORDER.get(x, 0), default=None)
    return best_b, best_a, best_t


def load_report_detail():
    """★보고서버전_최종.xlsx에서 분류체계 + 세부기술별 공표 데이터"""
    wb = openpyxl.load_workbook(REPORT_XLSX, data_only=True)

    # 1. 분류체계
    ws_cls = wb['분류체계']
    taxonomy = []
    cur_dae = None
    mid_num = 0
    prev_mid = None
    for r in range(2, ws_cls.max_row + 1):
        dae = ws_cls.cell(r, 1).value
        mid = ws_cls.cell(r, 2).value
        detail = ws_cls.cell(r, 3).value
        if dae:
            cur_dae = dae.strip()
        if mid and mid.strip() != prev_mid:
            prev_mid = mid.strip()
            mid_num += 1
        if detail:
            taxonomy.append({
                'dae': cur_dae, 'mid': mid.strip() if mid else prev_mid,
                'mid_num': mid_num, 'detail': detail.strip(),
            })

    # 2. 대분류 → 상위 중분류(감축/적응)
    ws_stat = wb['1.대분류별 통계(세부분류의 평균으로 계산)']
    tech_upper = {}
    for r in range(2, ws_stat.max_row + 1):
        upper = ws_stat.cell(r, 1).value
        tech = ws_stat.cell(r, 2).value
        if tech and upper:
            tech_upper[tech.strip()] = upper.strip()

    # 3. 세부분류별 통계
    ws_det = wb['4.세부분류별 통계']
    detail_stats = {}
    for r in range(2, ws_det.max_row + 1):
        dae = ws_det.cell(r, 1).value
        detail = ws_det.cell(r, 3).value
        country = ws_det.cell(r, 4).value
        score100 = ws_det.cell(r, 9).value
        year0 = ws_det.cell(r, 10).value
        if dae and detail and country:
            detail_stats[(dae.strip(), detail.strip(), country)] = {
                'score100': score100, 'year0': year0,
            }

    wb.close()
    return taxonomy, tech_upper, detail_stats


def load_research_l0():
    """L0 Raw에서 세부기술 × 국가별 연구역량/R&D 직접 계산"""
    wb = openpyxl.load_workbook(L0_RAW, data_only=True)
    ws = wb.active

    raw = defaultdict(lambda: defaultdict(list))
    for r in range(2, ws.max_row + 1):
        tech = ws.cell(r, 3).value
        detail = ws.cell(r, 4).value
        country = ws.cell(r, 5).value
        basic = ws.cell(r, 9).value
        app = ws.cell(r, 10).value
        trend = ws.cell(r, 11).value
        if not tech or not detail or not country:
            continue
        key = (_normalize_name(tech), _normalize_name(detail))
        raw[key][country].append({'basic': basic, 'app': app, 'trend': trend})
    wb.close()

    result = {}
    for key, country_data in raw.items():
        result[key] = {}
        for cn in ['한국', '중국', '일본', '미국']:
            recs = country_data.get(cn, [])
            if recs:
                b, a, t = _calc_research(recs)
                result[key][cn] = {'basic': b, 'app': a, 'trend': t}
        eu_b, eu_a, eu_t = _calc_eu(country_data)
        if eu_b is not None:
            result[key]['EU'] = {'basic': eu_b, 'app': eu_a, 'trend': eu_t}
    return result


def build_rows(taxonomy, tech_upper, detail_stats, research):
    """157건 세부기술 × 양식 38열 행 생성"""
    rows = []
    tech_num_counter = 0

    for t in taxonomy:
        tech_num_counter += 1
        dae = t['dae']
        dae_num = dae.split('.')[0].strip().zfill(2)

        upper = tech_upper.get(dae, '')
        classification = UPPER_TO_CLASS.get(upper)
        if not classification:
            classification = '감축' if int(dae_num) <= 26 else '적응'

        std_name = STD_38_NAMES.get(dae_num, dae)

        # 5개국 기술수준/격차
        levels = {}
        gaps = {}
        for cn in COUNTRIES:
            st = detail_stats.get((dae, t['detail'], cn))
            if st:
                levels[cn] = st['score100']
                gaps[cn] = st['year0']

        top_countries = [cn for cn in COUNTRIES if levels.get(cn) is not None
                         and abs(levels[cn] - 100) < 0.01]
        top_country = '·'.join(top_countries) if top_countries else None

        # 기술그룹 (100환산 비율 자동 산출)
        max_level = max((v for v in levels.values() if v is not None), default=0)
        def _group(lv):
            if lv is None or max_level == 0:
                return None
            ratio = lv / max_level
            if ratio >= 0.95:
                return '선도'
            elif ratio >= 0.80:
                return '추격'
            return '후발'

        r_key = (_normalize_name(dae), _normalize_name(t['detail']))
        res = research.get(r_key, {})

        row = [
            '2025', classification, t['mid_num'], t['mid'], std_name,
            tech_num_counter, t['detail'], top_country,
        ]

        for cn in COUNTRIES:
            row.extend([
                round(levels[cn], 1) if cn in levels else None,
                gaps.get(cn),
                _group(levels.get(cn)),
            ])

        for cn in COUNTRIES:
            r_data = res.get(cn, {})
            row.extend([r_data.get('trend'), r_data.get('basic'), r_data.get('app')])

        rows.append(row)
    return rows


def write_excel(rows):
    """2020 양식 기반으로 2025 데이터 엑셀 생성"""
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active

    ref_styles = []
    for c in range(1, 39):
        cell = ws.cell(2, c)
        ref_styles.append({
            'font': copy(cell.font), 'alignment': copy(cell.alignment),
            'border': copy(cell.border), 'number_format': cell.number_format,
        })

    old_max = ws.max_row
    for r in range(2, old_max + 1):
        for c in range(1, 39):
            ws.cell(r, c, None)
    extra = old_max - 1 - len(rows)
    if extra > 0:
        ws.delete_rows(len(rows) + 2, extra)

    for i, row_data in enumerate(rows):
        r = i + 2
        for c_idx, value in enumerate(row_data):
            cell = ws.cell(r, c_idx + 1, value)
            if c_idx < len(ref_styles):
                style = ref_styles[c_idx]
                cell.font = copy(style['font'])
                cell.alignment = copy(style['alignment'])
                cell.border = copy(style['border'])
                cell.number_format = style['number_format']

    ws.title = '기후기술수준조사_세부기술별 데이터'
    wb.save(OUT)
    wb.close()
    print(f"  {len(rows)}행 × 38열 → {OUT.name}")


def main():
    print("=" * 60)
    print("2025 기후기술수준조사 세부기술별 데이터 생성")
    print("=" * 60)

    print("\n[1] ★보고서버전_최종.xlsx → 분류체계 + 세부기술 통계...")
    taxonomy, tech_upper, detail_stats = load_report_detail()
    print(f"  분류체계: {len(taxonomy)}건, 세부통계: {len(detail_stats)}건")

    print("\n[2] L0 Raw → 세부기술별 연구역량/R&D...")
    research = load_research_l0()
    matched = sum(1 for t in taxonomy
                  if (_normalize_name(t['dae']), _normalize_name(t['detail'])) in research)
    print(f"  연구역량 매칭: {matched}/{len(taxonomy)}")

    print("\n[3] 양식 행 생성...")
    rows = build_rows(taxonomy, tech_upper, detail_stats, research)
    print(f"  생성: {len(rows)}행")
    empty = sum(1 for row in rows for v in row if v is None)
    print(f"  빈 셀: {empty}건")

    print("\n[4] 엑셀 생성...")
    write_excel(rows)
    print("\n완료!")


if __name__ == '__main__':
    main()
