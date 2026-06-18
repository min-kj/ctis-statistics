"""
2025 기후기술수준조사 소분류별 데이터 생성 스크립트

소스:
  - ★보고서버전_최종.xlsx (L3) → 기술수준(100점 환산), 격차(0년 환산)
  - 2차_최종DATA_(3차 포함).xlsx (L0) → 연구역량(기초/응용), R&D 활동 경향
    * 한국/중국/일본/미국: 해당국 개인응답 평균
    * EU: EU 회원국 중 국가별 평균의 최고국 값
  - 보고서 PDF (L4) → 기술그룹(선도/추격/후발) 공표 값

출력: 기후기술수준조사_소분류별 데이터_2025추가.xlsx (2020 원본 + 2025 추가)
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from pathlib import Path
from collections import defaultdict, Counter
import openpyxl
from copy import copy

DATA = Path("src/data")

def _find(pattern):
    results = list(DATA.glob(pattern))
    if not results:
        raise FileNotFoundError(f"'{pattern}' not found in {DATA}")
    return results[0]

TEMPLATE = _find("*소분류별*2020*")
OUT = DATA / TEMPLATE.name.replace("2020", "2025")
REPORT_XLSX = _find("*보고서버전_최종*")
L0_RAW = _find("2025 수준조사 Rawdata/*2차_최종DATA*")

COUNTRIES = ['한국', '중국', '일본', '미국', 'EU']
EU_MEMBER_COUNTRIES = {
    '독일', '프랑스', '영국', '이탈리아', '스페인', '네덜란드', '벨기에',
    '오스트리아', '스웨덴', '덴마크', '핀란드', '아일랜드', '노르웨이', '스위스',
    '룩셈부르크', '폴란드', '체코', '포르투갈', '에스토니아',
}
RAW_SCORE_MAP = {1: 20, 2: 40, 3: 60, 4: 80, 5: 100}
RAW_TREND_MAP = {1: '급하강', 2: '하강', 3: '유지', 4: '상승', 5: '급상승'}
TREND_ORDER = {'급상승': 5, '상승': 4, '유지': 3, '하강': 2, '급하강': 1}

# 38대 표준 대분류명 (플랫폼탑재용 기준)
STD_38_NAMES = {
    '01': '태양광 기술',       '02': '태양열 기술',       '03': '풍력 기술',
    '04': '해양에너지 기술',    '05': '수력 기술',        '06': '수열 기술',
    '07': '지열 기술',         '08': '바이오에너지 기술',   '09': '수소·암모니아 기술',
    '10': '석탄액화·가스화 기술', '11': '원자력 기술',       '12': '핵융합에너지 기술',
    '13': '수소 기술',         '14': '바이오매스 기술',    '15': '폐자원 기술',
    '16': '발전효율 기술',      '17': '산업효율 기술',      '18': '수송효율 기술',
    '19': '건물효율 기술',
    '20': '이산화탄소(CO2)포집, 저장, 활용 기술',
    '21': '메탄(CH)처리 기술',
    '22': '기타 온실가스 처리 및 대체 기술',
    '23': '탄소흡수원 기술',    '24': '전력 통합 기술',     '25': '열 통합 기술',
    '26': '전력-비전력 부문간 결합 기술',
    '27': '기후변화 감시 및 진단 기술',  '28': '기후변화 예측 기술',
    '29': '기후변화 영향 평가 기술',
    '30': '기후변화 취약성 및 위험성 평가 기술',
    '31': '건강부문 기술',      '32': '물 부문 기술',       '33': '국토연안 부문 기술',
    '34': '농축수산 부문 기술',   '35': '산림·생태계 부문 기술', '36': '산업·에너지 부문 기술',
    '37': '적응조치의 효과평가 기술', '38': '기후변화 적응기반 기술',
}

# 보고서 PDF에서 추출한 38대 기술그룹 (소분류 전용 — 세부기술은 수치 자동 산출)
REPORT_GROUPS = {
    '01': {'미국': '선도', '한국': '선도', '중국': '선도', '일본': '선도', 'EU': '선도'},
    '02': {'미국': '선도', '한국': '후발', '중국': '선도', '일본': '후발', 'EU': '선도'},
    '03': {'미국': '후발', '한국': '추격', '중국': '후발', '일본': '추격', 'EU': '선도'},
    '04': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '05': {'미국': '선도', '한국': '후발', '중국': '선도', '일본': '선도', 'EU': '선도'},
    '06': {'미국': '선도', '한국': '후발', '중국': '선도', '일본': '선도', 'EU': '후발'},
    '07': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '선도', 'EU': '선도'},
    '08': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '09': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '선도', 'EU': '후발'},
    '10': {'미국': '선도', '한국': '후발', '중국': '선도', '일본': '선도', 'EU': '선도'},
    '11': {'미국': '선도', '한국': '후발', '중국': '선도', '일본': '후발', 'EU': '후발'},
    '12': {'미국': '선도', '한국': '후발', '중국': '선도', '일본': '선도', 'EU': '선도'},
    '13': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '선도', 'EU': '선도'},
    '14': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '15': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '선도', 'EU': '선도'},
    '16': {'미국': '후발', '한국': '후발', '중국': '선도', '일본': '후발', 'EU': '후발'},
    '17': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '18': {'미국': '후발', '한국': '선도', '중국': '선도', '일본': '선도', 'EU': '선도'},
    '19': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '20': {'미국': '선도', '한국': '추격', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '21': {'미국': '선도', '한국': '후발', '중국': '선도', '일본': '후발', 'EU': '후발'},
    '22': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '선도', 'EU': '선도'},
    '23': {'미국': '선도', '한국': '추격', '중국': '후발', '일본': '후발', 'EU': '후발'},
    '24': {'미국': '선도', '한국': '후발', '중국': '선도', '일본': '후발', 'EU': '선도'},
    '25': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '선도', 'EU': '선도'},
    '26': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '27': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '28': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '29': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '30': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '31': {'미국': '선도', '한국': '선도', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '32': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '선도', 'EU': '선도'},
    '33': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '34': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '선도', 'EU': '선도'},
    '35': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '36': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '37': {'미국': '선도', '한국': '후발', '중국': '후발', '일본': '후발', 'EU': '선도'},
    '38': {'미국': '선도', '한국': '후발', '중국': '선도', '일본': '후발', 'EU': '선도'},
}

UPPER_TO_CLASS = {
    'I. 에너지 생산': '감축', 'II. 연원료 대체': '감축',
    'III. 에너지 효율': '감축', 'IV. 온실가스 처리': '감축',
    'V. 에너지 융복합': '감축',
    'Ⅵ. 기후변화 모니터링': '적응', 'Ⅶ. 기후영향 평가 및 진단': '적응',
    'Ⅷ. 피해관리 및 탄력성 제고': '적응', 'Ⅸ. 정책기술 분석 및 평가': '적응',
}


def _normalize_name(name):
    """대분류명 정규화: ·/?/・ 통일, 공백 정리"""
    if not name:
        return name
    for ch in ['?', '\uff65', '\u30fb']:
        name = name.replace(ch, '\u00b7')
    return name.strip()


def _calc_research(records):
    """개인응답 리스트 → (기초점수, 응용점수, R&D경향)"""
    basics = [RAW_SCORE_MAP[d['basic']] for d in records if d.get('basic') in RAW_SCORE_MAP]
    apps = [RAW_SCORE_MAP[d['app']] for d in records if d.get('app') in RAW_SCORE_MAP]
    trends = [d['trend'] for d in records if d.get('trend') in RAW_TREND_MAP]
    avg_b = round(sum(basics) / len(basics), 1) if basics else None
    avg_a = round(sum(apps) / len(apps), 1) if apps else None
    rd = RAW_TREND_MAP[Counter(trends).most_common(1)[0][0]] if trends else None
    return avg_b, avg_a, rd


def _calc_eu(country_data):
    """EU 회원국별 평균 → 각 지표별 최고국 값"""
    eu_avgs = {}
    for cn in EU_MEMBER_COUNTRIES:
        recs = country_data.get(cn, [])
        if not recs:
            continue
        b, a, t = _calc_research(recs)
        eu_avgs[cn] = {'basic': b, 'app': a, 'trend': t}
    if not eu_avgs:
        return None, None, None
    best_b = max((v['basic'] for v in eu_avgs.values() if v['basic']), default=None)
    best_a = max((v['app'] for v in eu_avgs.values() if v['app']), default=None)
    best_t = max((v['trend'] for v in eu_avgs.values() if v['trend']),
                 key=lambda x: TREND_ORDER.get(x, 0), default=None)
    return best_b, best_a, best_t


def load_report_data():
    """★보고서버전_최종.xlsx에서 38대 분류별 공표 데이터 로드"""
    wb = openpyxl.load_workbook(REPORT_XLSX, data_only=True)

    ws = wb['1.대분류별 통계(세부분류의 평균으로 계산)']
    tech_data = defaultdict(dict)
    tech_upper = {}

    for r in range(2, ws.max_row + 1):
        upper = ws.cell(r, 1).value
        tech = ws.cell(r, 2).value
        country = ws.cell(r, 3).value
        if not tech or not country:
            continue
        tech = tech.strip()
        if upper:
            tech_upper[tech] = upper.strip()
        tech_data[tech][country] = {
            'score100': ws.cell(r, 7).value,
            'year0': ws.cell(r, 8).value,
        }

    ws_cls = wb['분류체계']
    detail_counts = defaultdict(int)
    prev_dae = None
    for r in range(2, ws_cls.max_row + 1):
        dae = ws_cls.cell(r, 1).value
        if dae:
            prev_dae = dae.strip()
        detail_counts[prev_dae] += 1

    wb.close()
    return tech_data, tech_upper, detail_counts


def load_research_data():
    """L0 Raw에서 대분류 × 국가별 연구역량/R&D 직접 계산"""
    wb = openpyxl.load_workbook(L0_RAW, data_only=True)
    ws = wb.active

    raw = defaultdict(lambda: defaultdict(list))
    for r in range(2, ws.max_row + 1):
        tech = ws.cell(r, 3).value
        country = ws.cell(r, 5).value
        basic = ws.cell(r, 9).value
        app = ws.cell(r, 10).value
        trend = ws.cell(r, 11).value
        if not tech or not country:
            continue
        raw[_normalize_name(tech)][country].append({
            'basic': basic, 'app': app, 'trend': trend,
        })
    wb.close()

    result = {}
    for tech, countries_data in raw.items():
        result[tech] = {}
        for cn in ['한국', '중국', '일본', '미국']:
            recs = countries_data.get(cn, [])
            if recs:
                b, a, t = _calc_research(recs)
                result[tech][cn] = {'basic': b, 'app': a, 'trend': t}
        eu_b, eu_a, eu_t = _calc_eu(countries_data)
        if eu_b is not None:
            result[tech]['EU'] = {'basic': eu_b, 'app': eu_a, 'trend': eu_t}
    return result


def build_rows(tech_data, tech_upper, detail_counts, research_agg):
    """38대 분류별 양식 행 생성"""
    rows = []
    for tech_name in sorted(tech_data.keys()):
        d = tech_data[tech_name]
        tech_num = tech_name.split('.')[0].strip().zfill(2)

        upper = tech_upper.get(tech_name, '')
        classification = UPPER_TO_CLASS.get(upper)
        if not classification:
            classification = '감축' if int(tech_num) <= 26 else '적응'

        top_countries = [c for c in COUNTRIES if c in d and d[c]['score100'] is not None
                         and abs(d[c]['score100'] - 100) < 0.01]
        top_country = '·'.join(top_countries) if top_countries else None

        detail_count = detail_counts.get(tech_name, detail_counts.get(tech_name.rstrip(), None))
        groups = REPORT_GROUPS.get(tech_num, {})
        std_name = STD_38_NAMES.get(tech_num, tech_name)

        row = ['2025', classification, std_name, top_country, detail_count]

        for c in COUNTRIES:
            cd = d.get(c, {})
            score = cd.get('score100')
            row.extend([
                round(score, 1) if score is not None else None,
                cd.get('year0'),
                groups.get(c),
            ])

        research = research_agg.get(_normalize_name(tech_name), {})
        for c in COUNTRIES:
            r_data = research.get(c, {})
            row.extend([r_data.get('trend'), r_data.get('basic'), r_data.get('app')])

        rows.append(row)
    return rows


def write_to_excel(rows):
    """기존 양식에 2025 데이터 추가"""
    import shutil
    shutil.copy2(TEMPLATE, OUT)
    wb = openpyxl.load_workbook(OUT)
    ws = wb.active

    last_row = ws.max_row
    print(f"  기존: {last_row}행 (헤더+2020 44행, 원본 이름 유지)")

    ref_styles = []
    for c in range(1, 36):
        cell = ws.cell(2, c)
        ref_styles.append({
            'font': copy(cell.font),
            'alignment': copy(cell.alignment),
            'border': copy(cell.border),
            'number_format': cell.number_format,
        })

    start_row = last_row + 1
    for i, row_data in enumerate(rows):
        r = start_row + i
        for c_idx, value in enumerate(row_data):
            cell = ws.cell(r, c_idx + 1, value)
            if c_idx < len(ref_styles):
                style = ref_styles[c_idx]
                cell.font = copy(style['font'])
                cell.alignment = copy(style['alignment'])
                cell.border = copy(style['border'])
                cell.number_format = style['number_format']

    wb.save(OUT)
    wb.close()
    print(f"  추가: 2025 {len(rows)}행 → 전체 {start_row + len(rows) - 1}행")


def main():
    print("=" * 60)
    print("2025 기후기술수준조사 소분류별 데이터 생성")
    print("=" * 60)

    print("\n[1] ★보고서버전_최종.xlsx 로드...")
    tech_data, tech_upper, detail_counts = load_report_data()
    print(f"  38대 분류: {len(tech_data)}건")

    print("\n[2] L0 Raw → 연구역량/R&D 집계...")
    research_agg = load_research_data()
    matched = sum(1 for t in tech_data if _normalize_name(t) in research_agg)
    print(f"  연구역량 매칭: {matched}/{len(tech_data)}")

    print("\n[3] 양식 행 생성...")
    rows = build_rows(tech_data, tech_upper, detail_counts, research_agg)
    print(f"  생성: {len(rows)}행")

    print("\n[4] 엑셀 파일 생성...")
    write_to_excel(rows)
    print(f"\n  → {OUT.name}")


if __name__ == '__main__':
    main()
