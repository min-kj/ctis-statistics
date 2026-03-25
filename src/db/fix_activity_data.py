"""
Fix activity survey data: re-parse from correct WT/TAB tables with per-category
per-org_size breakdown.

The previous loader only extracted '전체' values per category and total-level
org_size breakdown. This script:
1. Clears all activity_survey data
2. Re-parses from the correct 규모별 tables (A3_1_3/A3_1_4, B2_1_3/B2_1_4,
   C1 규모별) with per-category × per-org_size data
3. Also re-parses the non-규모별 tables (A3_1_1, B2_1_1, C1 simple) for the
   overall '전체' values
4. Re-parses A3-1 매출 미발생 이유
5. Verifies against known homepage values
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
import re
import sqlite3
import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # src/
DB_PATH = os.path.join(BASE, 'db', 'ctis_stats.db')
DATA_DIR = os.path.join(BASE, 'data', '활동조사')

FILE_22 = os.path.join(DATA_DIR, '기후기술조사(WT테이블)0207 -22년도.xlsx')
FILE_23 = os.path.join(DATA_DIR, '2023년도 TAB_기후기술_V1_가중.xlsx')

# ── Helpers ─────────────────────────────────────────────────────────────

def normalize_name(name):
    s = str(name).strip()
    s = re.sub(r'^\(감축\)', '', s)
    s = re.sub(r'^\(적응\)', '', s)
    s = re.sub(r'^\d{1,2}\.\s*', '', s)
    s = s.replace('·', '.')
    s = s.strip()
    return s


def load_category_map(conn):
    cur = conn.cursor()
    cur.execute("SELECT category_id, category_name FROM tech_category WHERE taxonomy_id=3")
    mapping = {}
    for cid, cname in cur.fetchall():
        mapping[normalize_name(cname)] = cid
    return mapping


def _to_float(val):
    if val is None or val == '.' or val == '-' or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def find_table_row(ws, marker_text):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        val = str(cell.value) if cell.value else ''
        if marker_text in val:
            return cell.row
    return None


# ── Org-size column maps ───────────────────────────────────────────────

ORG_SIZE_LABELS = ['100억미만', '100~500억', '500~1000억', '1000~2000억', '2000억이상', '공공기관등']

# For 22년도 WT 규모별 tables: header has pairs (평균, 합계) for each org_size
# Columns (1-based): 3=사례수, 4=전체평균, 5=전체합계,
#   6=층1평균, 7=층1합계, 8=층2평균, 9=층2합계,
#   10=층3평균, 11=층3합계, 12=층4평균, 13=층4합계,
#   14=층5평균, 15=층5합계, 16=기타평균, 17=기타합계
ORG_SIZE_COLS_22 = {
    '전체': 5,
    '100억미만': 7,
    '100~500억': 9,
    '500~1000억': 11,
    '1000~2000억': 13,
    '2000억이상': 15,
    '공공기관등': 17,
}

# For 23년도 TAB tables: single column per org_size (합계 only)
# Columns (1-based): 3=사례수, 4=[합계](전체),
#   5=100억미만, 6=100~500억, 7=500~1000억, 8=1000~2000억, 9=2000억이상, 10=기타
ORG_SIZE_COLS_23 = {
    '전체': 4,
    '100억미만': 5,
    '100~500억': 6,
    '500~1000억': 7,
    '1000~2000억': 8,
    '2000억이상': 9,
    '공공기관등': 10,
}

# For org_size row matching in bottom section of tables
ORG_SIZE_ROW_PATTERNS_22 = [
    (r'층1|100억 미만', '100억미만'),
    (r'층2|100억 이상 500억 미만', '100~500억'),
    (r'층3|500억 이상 1천억 미만', '500~1000억'),
    (r'층4|1천억 이상 2천억 미만', '1000~2000억'),
    (r'층5|2천억 이상', '2000억이상'),
    (r'매출액 정보 없음|공공기관 등', '공공기관등'),
]

ORG_SIZE_ROW_PATTERNS_23 = [
    (r'1\.\s*100억 미만', '100억미만'),
    (r'2\.\s*100억 이상', '100~500억'),
    (r'3\.\s*500억 이상', '500~1000억'),
    (r'4\.\s*1000억 이상', '1000~2000억'),
    (r'5\.\s*2000억 이상', '2000억이상'),
    (r'6\.\s*기타 공공부문|기타 공공부문 등', '공공기관등'),
]


def match_org_size_label(text, patterns):
    if not text:
        return None
    s = str(text).strip()
    for pat, label in patterns:
        if re.search(pat, s):
            return label
    return None


# ── Parser for 규모별 tables (per-category × per-org_size) ────────────

def parse_table_22_규모별(ws, marker_row, cat_map, org_size_cols):
    """
    Parse a 22년도 규모별 table.
    Returns list of (category_id_or_None, org_size, value).
    Extracts 합계 column for each org_size for each category row.
    """
    results = []
    r = marker_row + 4  # [전  체] row

    # Read [전체] row - all org_size values
    for org_size, col in org_size_cols.items():
        val = ws.cell(r, col).value
        fval = _to_float(val)
        if fval is not None:
            results.append((None, org_size, fval))

    r += 1  # Skip 감축 subtotal
    r += 1  # Skip 적응 subtotal
    r += 1  # First category row ('기후기술 분야' | category)

    # Read category rows
    while r <= marker_row + 40:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value

        if col_a and '기관규모' in str(col_a):
            break
        if col_a is None and col_b is None:
            break

        cat_name_raw = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        if not cat_name_raw:
            r += 1
            continue

        norm = normalize_name(cat_name_raw)
        cat_id = cat_map.get(norm)
        if cat_id is None:
            print(f"  WARNING: No category match for '{cat_name_raw}' (normalized: '{norm}')")
            r += 1
            continue

        for org_size, col in org_size_cols.items():
            val = ws.cell(r, col).value
            fval = _to_float(val)
            if fval is not None:
                results.append((cat_id, org_size, fval))

        r += 1

    return results


def parse_table_23_규모별(ws, marker_row, cat_map, org_size_cols):
    """
    Parse a 23년도 규모별 table.
    Returns list of (category_id_or_None, org_size, value).
    """
    results = []
    r = marker_row + 3  # [전체] row

    # Read [전체] row
    for org_size, col in org_size_cols.items():
        val = ws.cell(r, col).value
        fval = _to_float(val)
        if fval is not None:
            results.append((None, org_size, fval))

    r += 1  # 감축 subtotal
    r += 1  # 적응 subtotal
    r += 1  # first category row

    while r <= marker_row + 40:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value

        if col_a and '기관규모' in str(col_a):
            break
        if col_a is None and col_b is None:
            break

        cat_name_raw = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        if not cat_name_raw:
            r += 1
            continue

        # Skip section headers like "기후기술 분야(감축)"
        if '기후기술 분야' in cat_name_raw and col_b is None:
            r += 1
            continue

        norm = normalize_name(cat_name_raw)
        cat_id = cat_map.get(norm)
        if cat_id is None:
            print(f"  WARNING: No category match for '{cat_name_raw}' (normalized: '{norm}')")
            r += 1
            continue

        for org_size, col in org_size_cols.items():
            val = ws.cell(r, col).value
            fval = _to_float(val)
            if fval is not None:
                results.append((cat_id, org_size, fval))

        r += 1

    return results


# ── Parser for simple tables (C1 전체/기후기술 without 규모별) ──────────

def parse_table_22_simple(ws, marker_row, cat_map, value_col):
    """Parse 22년도 simple table (no 규모별 columns). Returns per-category '전체' values."""
    results = []
    r = marker_row + 4  # [전  체] row

    val = ws.cell(r, value_col).value
    fval = _to_float(val)
    if fval is not None:
        results.append((None, '전체', fval))

    r += 3  # skip subtotals, go to first category

    while r <= marker_row + 40:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value

        if col_a and '기관규모' in str(col_a):
            break
        if col_a is None and col_b is None:
            break

        cat_name_raw = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        if not cat_name_raw:
            r += 1
            continue

        norm = normalize_name(cat_name_raw)
        cat_id = cat_map.get(norm)
        if cat_id is None:
            print(f"  WARNING: No category match for '{cat_name_raw}' (normalized: '{norm}')")
            r += 1
            continue

        val = ws.cell(r, value_col).value
        fval = _to_float(val)
        if fval is not None:
            results.append((cat_id, '전체', fval))
        r += 1

    return results


def parse_table_23_simple(ws, marker_row, cat_map, value_col):
    """Parse 23년도 simple table. Returns per-category '전체' values."""
    results = []
    r = marker_row + 3  # [전체] row

    val = ws.cell(r, value_col).value
    fval = _to_float(val)
    if fval is not None:
        results.append((None, '전체', fval))

    r += 3  # skip subtotals

    while r <= marker_row + 40:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value

        if col_a and '기관규모' in str(col_a):
            break
        if col_a is None and col_b is None:
            break

        cat_name_raw = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        if not cat_name_raw:
            r += 1
            continue
        if '기후기술 분야' in cat_name_raw and col_b is None:
            r += 1
            continue

        norm = normalize_name(cat_name_raw)
        cat_id = cat_map.get(norm)
        if cat_id is None:
            print(f"  WARNING: No category match for '{cat_name_raw}' (normalized: '{norm}')")
            r += 1
            continue

        val = ws.cell(r, value_col).value
        fval = _to_float(val)
        if fval is not None:
            results.append((cat_id, '전체', fval))
        r += 1

    return results


# ── Parser for 매출미발생 사유 (A3-1) ──────────────────────────────────

REASON_COLS = [
    (4, '사업화자금부족'),
    (5, '시장여건변화대응미흡'),
    (6, '제품서비스완성도미비'),
    (7, '판로개척실패'),
    (8, '추가기술개발실패'),
    (9, '사업화전문인력부족'),
    (10, '마케팅홍보역량부족'),
    (11, '법규제정보획득대응어려움'),
    (12, '기타'),
]


def parse_reason_table(ws, marker_row, cat_map, org_patterns, has_go_row):
    """
    Parse 매출미발생이유 table.
    Returns list of (category_id_or_None, org_size, reason_name, value).
    """
    results = []

    # Find [전체] row
    r = marker_row + 1
    for scan_r in range(marker_row + 1, marker_row + 6):
        val = str(ws.cell(scan_r, 1).value or '')
        if '전' in val and '체' in val:
            r = scan_r
            break

    # Read [전체]
    for col_idx, reason_name in REASON_COLS:
        val = ws.cell(r, col_idx).value
        fval = _to_float(val)
        if fval is not None:
            results.append((None, '전체', reason_name, fval))
    r += 1

    # Skip subtotals
    r += 1  # 감축
    r += 1  # 적응

    # Category rows
    while r <= marker_row + 45:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value

        if col_a and '기관규모' in str(col_a):
            break
        if col_a is None and col_b is None:
            break

        cat_name_raw = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        if not cat_name_raw:
            r += 1
            continue
        if '기후기술 분야' in cat_name_raw and col_b is None:
            r += 1
            continue

        norm = normalize_name(cat_name_raw)
        cat_id = cat_map.get(norm)
        if cat_id is None:
            print(f"  WARNING (a2): No match for '{cat_name_raw}' (normalized: '{norm}')")
            r += 1
            continue

        for col_idx, reason_name in REASON_COLS:
            val = ws.cell(r, col_idx).value
            fval = _to_float(val)
            if fval is not None:
                results.append((cat_id, '전체', reason_name, fval))
        r += 1

    # Org-size rows
    while r <= marker_row + 55:
        col_a = ws.cell(r, 1).value
        col_b = ws.cell(r, 2).value
        if col_a is None and col_b is None:
            break

        label = str(col_b).strip() if col_b else (str(col_a).strip() if col_a else '')
        org = match_org_size_label(label, org_patterns)
        if org:
            for col_idx, reason_name in REASON_COLS:
                val = ws.cell(r, col_idx).value
                fval = _to_float(val)
                if fval is not None:
                    results.append((None, org, reason_name, fval))
        r += 1

    return results


# ── Main ────────────────────────────────────────────────────────────────

def main():
    conn = sqlite3.connect(DB_PATH)
    cat_map = load_category_map(conn)
    print(f"Loaded {len(cat_map)} T22 categories")

    cur = conn.cursor()

    # ── Snapshot before ──
    print("\n=== BEFORE: Current data snapshot ===")
    cur.execute("""
        SELECT survey_year, metric_type, org_size, COUNT(*), SUM(metric_value)
        FROM activity_survey
        GROUP BY survey_year, metric_type, org_size
        ORDER BY survey_year, metric_type, org_size
    """)
    before_summary = cur.fetchall()
    for row in before_summary:
        yr, mt, org, cnt, total = row
        print(f"  {yr} | {mt:40s} | {org:12s} | {cnt:3d} rows | {total:>18,.2f}" if total else
              f"  {yr} | {mt:40s} | {org:12s} | {cnt:3d} rows | N/A")

    # Check specific before values
    cur.execute("SELECT metric_value FROM activity_survey WHERE metric_type='revenue_climate' AND category_id=83 AND org_size='전체' AND survey_year=2022")
    row = cur.fetchone()
    before_solar_rc = row[0] if row else None
    print(f"\n  Before 2022 태양광.열 revenue_climate 전체: {before_solar_rc}")

    cur.execute("SELECT SUM(metric_value) FROM activity_survey WHERE metric_type='revenue_climate' AND category_id IS NOT NULL AND org_size='전체' AND survey_year=2022")
    row = cur.fetchone()
    before_total_rc = row[0] if row else None
    print(f"  Before 2022 revenue_climate 전체 SUM(categories): {before_total_rc}")

    # ── Clear ──
    cur.execute("SELECT COUNT(*) FROM activity_survey")
    old_count = cur.fetchone()[0]
    cur.execute("DELETE FROM activity_survey")
    conn.commit()
    print(f"\n=== Deleted {old_count} rows ===")

    # ── Load workbooks ──
    print(f"\nLoading 22년도: {os.path.basename(FILE_22)}")
    wb22 = openpyxl.load_workbook(FILE_22, data_only=True)
    ws22 = wb22['테이블']

    print(f"Loading 23년도: {os.path.basename(FILE_23)}")
    wb23 = openpyxl.load_workbook(FILE_23, data_only=True)
    ws23 = wb23['Table']

    all_inserts = []

    # ──────────────────────────────────────────────────────────────────
    # 2022 data
    # ──────────────────────────────────────────────────────────────────
    print("\n=== Parsing 2022 data ===")

    # --- revenue (전체 매출액) ---
    # A3_1_3: 규모별 전체 매출액 → per-category × per-org_size
    row = find_table_row(ws22, 'A3_1_3. 2022년 기업 규모별 매출 현황_전체 매출액')
    if row:
        data = parse_table_22_규모별(ws22, row, cat_map, ORG_SIZE_COLS_22)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'revenue', val, '백만원', 1))
        print(f"  revenue (A3_1_3 규모별): {len(data)} rows from row {row}")
    else:
        print("  WARNING: A3_1_3 not found!")

    # --- revenue_climate (기후기술 매출액) ---
    # A3_1_4: 규모별 기후기술 매출액 → per-category × per-org_size
    row = find_table_row(ws22, 'A3_1_4. 2022년 기업 규모별 매출 현황_기후기술 분야 매출액')
    if row:
        data = parse_table_22_규모별(ws22, row, cat_map, ORG_SIZE_COLS_22)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'revenue_climate', val, '백만원', 1))
        print(f"  revenue_climate (A3_1_4 규모별): {len(data)} rows from row {row}")
    else:
        print("  WARNING: A3_1_4 not found!")

    # --- revenue_no_reason (매출 미발생 사유) ---
    row = find_table_row(ws22, 'A3-1. 매출 미발생')
    if row:
        data = parse_reason_table(ws22, row, cat_map, ORG_SIZE_ROW_PATTERNS_22, has_go_row=True)
        for cat_id, org, reason, val in data:
            mt = f'revenue_no_reason_{reason}'
            all_inserts.append((2022, cat_id, org, mt, val, '%', 1))
        print(f"  revenue_no_reason (A3-1): {len(data)} rows from row {row}")

    # --- rnd_expense (전체 연구개발비) ---
    # B2_1_3: 규모별 전체 연구개발비
    row = find_table_row(ws22, 'B2_1_3. 2022년 기업 규모별 연구개발 현황_전체 연구개발비')
    if row:
        data = parse_table_22_규모별(ws22, row, cat_map, ORG_SIZE_COLS_22)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'rnd_expense', val, '백만원', 1))
        print(f"  rnd_expense (B2_1_3 규모별): {len(data)} rows from row {row}")
    else:
        print("  WARNING: B2_1_3 not found!")

    # --- rnd_expense_climate (기후기술 연구개발비) ---
    # B2_1_4: 규모별 기후기술 연구개발비
    row = find_table_row(ws22, 'B2_1_4. 2022년 기업 규모별 연구개발 현황_기후기술 분야 연구개발비')
    if row:
        data = parse_table_22_규모별(ws22, row, cat_map, ORG_SIZE_COLS_22)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'rnd_expense_climate', val, '백만원', 1))
        print(f"  rnd_expense_climate (B2_1_4 규모별): {len(data)} rows from row {row}")
    else:
        print("  WARNING: B2_1_4 not found!")

    # --- employee_count (전체 종사자수) ---
    # C1 기업 규모별 종업원 수_전체: 전체
    row = find_table_row(ws22, 'C1. 기업 규모별 종업원 수_전체 : 전체')
    if row:
        data = parse_table_22_규모별(ws22, row, cat_map, ORG_SIZE_COLS_22)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'employee_count', val, '명', 1))
        print(f"  employee_count (C1 규모별 전체): {len(data)} rows from row {row}")
    else:
        # Fallback to simple C1 table
        row = find_table_row(ws22, 'C1. 종업원 수_전체 종사자수')
        if row:
            data = parse_table_22_simple(ws22, row, cat_map, 5)
            for cat_id, org, val in data:
                all_inserts.append((2022, cat_id, org, 'employee_count', val, '명', 1))
            print(f"  employee_count (C1 simple): {len(data)} rows from row {row}")

    # --- employee_count_climate (기후기술 종사자수) ---
    # C1 기업 규모별 종업원 수_기후기술 분야: 전체
    row = find_table_row(ws22, 'C1. 기업 규모별 종업원 수_기후기술 분야 : 전체')
    if row:
        data = parse_table_22_규모별(ws22, row, cat_map, ORG_SIZE_COLS_22)
        for cat_id, org, val in data:
            all_inserts.append((2022, cat_id, org, 'employee_count_climate', val, '명', 1))
        print(f"  employee_count_climate (C1 규모별 기후기술): {len(data)} rows from row {row}")
    else:
        row = find_table_row(ws22, 'C1. 종업원 수_기후기술 분야 종사자 수')
        if row:
            data = parse_table_22_simple(ws22, row, cat_map, 5)
            for cat_id, org, val in data:
                all_inserts.append((2022, cat_id, org, 'employee_count_climate', val, '명', 1))
            print(f"  employee_count_climate (C1 simple): {len(data)} rows from row {row}")

    # ──────────────────────────────────────────────────────────────────
    # 2023 data
    # ──────────────────────────────────────────────────────────────────
    print("\n=== Parsing 2023 data ===")

    # --- revenue (전체 매출액) ---
    # A3_1_1: 전체 매출액 (2023) - already has org_size columns
    row = find_table_row(ws23, 'A3_1_1. 전체 매출액(2023년')
    if row:
        data = parse_table_23_규모별(ws23, row, cat_map, ORG_SIZE_COLS_23)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'revenue', val, '백만원', 1))
        print(f"  revenue (A3_1_1): {len(data)} rows from row {row}")

    # --- revenue_climate (기후기술 매출액) ---
    # A3_1_3: 기후기술분야 매출액 (2023)
    row = find_table_row(ws23, 'A3_1_3. 기후기술분야 매출액(2023년')
    if row:
        data = parse_table_23_규모별(ws23, row, cat_map, ORG_SIZE_COLS_23)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'revenue_climate', val, '백만원', 1))
        print(f"  revenue_climate (A3_1_3): {len(data)} rows from row {row}")

    # --- revenue_no_reason ---
    row = find_table_row(ws23, 'A3-1. 매출 미발생')
    if row:
        data = parse_reason_table(ws23, row, cat_map, ORG_SIZE_ROW_PATTERNS_23, has_go_row=False)
        for cat_id, org, reason, val in data:
            mt = f'revenue_no_reason_{reason}'
            all_inserts.append((2023, cat_id, org, mt, val, '%', 1))
        print(f"  revenue_no_reason (A3-1): {len(data)} rows from row {row}")

    # --- rnd_expense (전체 연구개발비) ---
    row = find_table_row(ws23, 'B2_1_1. 전체 연구개발비 (2023년')
    if row:
        data = parse_table_23_규모별(ws23, row, cat_map, ORG_SIZE_COLS_23)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'rnd_expense', val, '백만원', 1))
        print(f"  rnd_expense (B2_1_1): {len(data)} rows from row {row}")

    # --- rnd_expense_climate (기후기술 연구개발비) ---
    row = find_table_row(ws23, 'B2_1_3. 기후기술분야 연구개발비(2023년')
    if row:
        data = parse_table_23_규모별(ws23, row, cat_map, ORG_SIZE_COLS_23)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'rnd_expense_climate', val, '백만원', 1))
        print(f"  rnd_expense_climate (B2_1_3): {len(data)} rows from row {row}")

    # --- employee_count (전체 종사자수) ---
    row = find_table_row(ws23, 'C1_1_1. 전체 종사자 수(전체)')
    if row:
        data = parse_table_23_규모별(ws23, row, cat_map, ORG_SIZE_COLS_23)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'employee_count', val, '명', 1))
        print(f"  employee_count (C1_1_1): {len(data)} rows from row {row}")

    # --- employee_count_climate (기후기술 종사자수) ---
    row = find_table_row(ws23, 'C1_2_1. 기후기술 분야별 종업원 수(전체)')
    if row:
        data = parse_table_23_규모별(ws23, row, cat_map, ORG_SIZE_COLS_23)
        for cat_id, org, val in data:
            all_inserts.append((2023, cat_id, org, 'employee_count_climate', val, '명', 1))
        print(f"  employee_count_climate (C1_2_1): {len(data)} rows from row {row}")

    # ── Insert all ──
    print(f"\n=== Inserting {len(all_inserts)} rows ===")
    cur.executemany(
        """INSERT INTO activity_survey
           (survey_year, category_id, org_size, metric_type, metric_value, metric_unit, is_weighted)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        all_inserts
    )
    conn.commit()

    # ── After summary ──
    print("\n=== AFTER: New data summary ===")
    cur.execute("""
        SELECT survey_year, metric_type, org_size, COUNT(*), SUM(metric_value)
        FROM activity_survey
        GROUP BY survey_year, metric_type, org_size
        ORDER BY survey_year, metric_type, org_size
    """)
    for row in cur.fetchall():
        yr, mt, org, cnt, total = row
        print(f"  {yr} | {mt:40s} | {org:12s} | {cnt:3d} rows | {total:>18,.2f}" if total else
              f"  {yr} | {mt:40s} | {org:12s} | {cnt:3d} rows | N/A")

    cur.execute("SELECT COUNT(*) FROM activity_survey")
    total_rows = cur.fetchone()[0]
    print(f"\nTotal rows: {total_rows}")

    # ── Verification ──
    print("\n=== VERIFICATION ===")

    # Check 1: 2022 태양광.열 revenue_climate 전체
    cur.execute("SELECT metric_value FROM activity_survey WHERE metric_type='revenue_climate' AND category_id=83 AND org_size='전체' AND survey_year=2022")
    row = cur.fetchone()
    val = row[0] if row else None
    expected = 30404081
    status = "OK" if val and abs(val - expected) < 10 else "MISMATCH"
    print(f"  2022 태양광.열 revenue_climate 전체: {val:,.2f} (expected ~{expected:,}) [{status}]")

    # Check 2: 2022 전체 revenue_climate - sum of 6 org_size groups
    cur.execute("""
        SELECT SUM(metric_value) FROM activity_survey
        WHERE metric_type='revenue_climate' AND category_id IS NULL
        AND org_size != '전체' AND survey_year=2022
    """)
    row = cur.fetchone()
    val = row[0] if row else None
    print(f"  2022 전체 revenue_climate SUM(6 org_sizes): {val:,.2f}" if val else "  N/A")

    # Check 3: Sum of all category 전체 values
    cur.execute("""
        SELECT SUM(metric_value) FROM activity_survey
        WHERE metric_type='revenue_climate' AND category_id IS NOT NULL
        AND org_size='전체' AND survey_year=2022
    """)
    row = cur.fetchone()
    val = row[0] if row else None
    print(f"  2022 revenue_climate SUM(22 categories, 전체): {val:,.2f}" if val else "  N/A")

    # Check 4: Per-category org_size breakdown count
    cur.execute("""
        SELECT COUNT(DISTINCT category_id || '-' || org_size)
        FROM activity_survey
        WHERE metric_type='revenue_climate' AND category_id IS NOT NULL
        AND org_size != '전체' AND survey_year=2022
    """)
    row = cur.fetchone()
    print(f"  2022 revenue_climate per-category per-org_size entries: {row[0]}")

    # Check 5: 2022 태양광.열 per org_size
    cur.execute("""
        SELECT org_size, metric_value FROM activity_survey
        WHERE metric_type='revenue_climate' AND category_id=83 AND survey_year=2022
        ORDER BY org_size
    """)
    print("  2022 태양광.열 revenue_climate by org_size:")
    for org, val in cur.fetchall():
        print(f"    {org:12s}: {val:>18,.2f}")

    # Check 6: 2023 values
    cur.execute("SELECT metric_value FROM activity_survey WHERE metric_type='revenue_climate' AND category_id=83 AND org_size='전체' AND survey_year=2023")
    row = cur.fetchone()
    val = row[0] if row else None
    print(f"\n  2023 태양광.열 revenue_climate 전체: {val:,.2f}" if val else "  N/A")

    # Check 7: revenue_no_reason counts
    cur.execute("""
        SELECT survey_year, COUNT(DISTINCT category_id)
        FROM activity_survey
        WHERE metric_type LIKE 'revenue_no_reason_%' AND org_size='전체'
        GROUP BY survey_year
    """)
    for yr, cnt in cur.fetchall():
        print(f"  {yr} revenue_no_reason categories: {cnt}")

    conn.close()
    print("\nDone!")


if __name__ == '__main__':
    main()
